from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

from apps.reportes.dtos import ReportResult


def export_xlsx(report: ReportResult) -> HttpResponse:
    """
    Genera un XLSX simple con cabecera y filas usando openpyxl.
    SRP: solo renderizado a XLSX a partir de ReportResult.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(report.columns))
    title_cell = ws.cell(row=1, column=1, value=report.title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="left")

    # Filtros (opcional)
    if report.filters_summary:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(report.columns))
        filters_text = " | ".join([f"{k}: {v}" for k, v in report.filters_summary.items() if v not in [None, ""]])
        ws.cell(row=2, column=1, value=filters_text)

    # Cabecera
    header_row = 4
    for col_idx, col_name in enumerate(report.columns, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=col_name)
        c.font = Font(bold=True)

    # Filas
    for row_idx, row in enumerate(report.rows, start=header_row + 1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto ancho
    for col_idx in range(1, len(report.columns) + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 18

    stream = BytesIO()
    wb.save(stream)
    xlsx_value = stream.getvalue()
    stream.close()

    response = HttpResponse(
        xlsx_value, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{report.title}.xlsx"'
    return response


