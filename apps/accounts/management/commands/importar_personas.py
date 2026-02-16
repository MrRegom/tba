"""
Comando de Django para importar personas desde archivo Excel.

Uso:
    python manage.py importar_personas

El comando lee el archivo personas.xlsx de la raíz del proyecto e importa:
- Usuarios en auth_user (username = documento_identidad)
- Personas en tba_persona (vinculadas a auth_user)
"""
from django.core.management.base import BaseCommand
from django.db import transaction
from django.contrib.auth import get_user_model
from openpyxl import load_workbook
from datetime import date
from apps.accounts.models import Persona

User = get_user_model()

# Contraseña hasheada que se usará para todos los usuarios
PASSWORD_HASHEADA = 'pbkdf2_sha256$1000000$ab0lVLlIKc9oT68SZYCFM5$VuBXxbxyJ3I7aHA1y2qN/mrM9p8g0lzce9XXDrz8LzY='

# Fecha por defecto para nacimiento (no viene en el Excel)
FECHA_NACIMIENTO_DEFAULT = date(1990, 1, 1)


class Command(BaseCommand):
    help = 'Importa personas desde archivo personas.xlsx'

    def add_arguments(self, parser):
        parser.add_argument(
            '--archivo',
            type=str,
            default='personas.xlsx',
            help='Ruta al archivo Excel (default: personas.xlsx en raíz del proyecto)'
        )
        parser.add_argument(
            '--actualizar',
            action='store_true',
            help='Actualizar registros existentes si ya existen'
        )

    def handle(self, *args, **options):
        archivo = options['archivo']
        actualizar = options['actualizar']

        self.stdout.write(self.style.SUCCESS(f'Iniciando importación desde {archivo}...'))

        try:
            wb = load_workbook(archivo)
            ws = wb.active

            # Obtener encabezados
            headers = [cell.value for cell in ws[1]]

            # Verificar que tenemos las columnas esperadas
            columnas_requeridas = ['documento_identidad', 'nombres', 'apellido1', 'apellido2', 'sexo', 'correo']
            for col in columnas_requeridas:
                if col not in headers:
                    raise ValueError(f'Columna requerida no encontrada: {col}')

            # Obtener índices de columnas
            idx_doc = headers.index('documento_identidad')
            idx_nombres = headers.index('nombres')
            idx_apellido1 = headers.index('apellido1')
            idx_apellido2 = headers.index('apellido2')
            idx_sexo = headers.index('sexo')
            idx_fecha_nac = headers.index('fecha_nacimiento') if 'fecha_nacimiento' in headers else None
            idx_talla = headers.index('talla') if 'talla' in headers else None
            idx_zapato = headers.index('numero_zapato') if 'numero_zapato' in headers else None
            idx_correo = headers.index('correo')

            total_filas = ws.max_row - 1  # Excluir encabezado
            procesadas = 0
            creadas = 0
            actualizadas = 0
            errores = 0

            self.stdout.write(f'Total de filas a procesar: {total_filas}')

            # Procesar cada fila
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Extraer datos de la fila
                    documento_identidad = str(row[idx_doc]).strip() if row[idx_doc] else None

                    if not documento_identidad:
                        self.stdout.write(
                            self.style.WARNING(f'Fila {row_num}: documento_identidad vacío, omitiendo...')
                        )
                        continue

                    nombres = str(row[idx_nombres]).strip() if row[idx_nombres] else ''
                    apellido1 = str(row[idx_apellido1]).strip() if row[idx_apellido1] else ''
                    apellido2 = str(row[idx_apellido2]).strip() if row[idx_apellido2] else ''
                    sexo_excel = row[idx_sexo]
                    correo = str(row[idx_correo]).strip() if row[idx_correo] else ''

                    # Convertir sexo de Excel (1=F, 2=M) a modelo
                    if sexo_excel == 1:
                        sexo = 'F'
                    elif sexo_excel == 2:
                        sexo = 'M'
                    else:
                        sexo = 'O'

                    # Fecha de nacimiento (usar default si no viene en Excel)
                    if idx_fecha_nac and row[idx_fecha_nac]:
                        fecha_nacimiento = row[idx_fecha_nac]
                        if isinstance(fecha_nacimiento, str):
                            from datetime import datetime
                            fecha_nacimiento = datetime.strptime(fecha_nacimiento, '%Y-%m-%d').date()
                    else:
                        fecha_nacimiento = FECHA_NACIMIENTO_DEFAULT

                    # Campos opcionales
                    talla = str(row[idx_talla]).strip() if idx_talla and row[idx_talla] else None
                    numero_zapato = str(row[idx_zapato]).strip() if idx_zapato and row[idx_zapato] else None

                    # Procesar con transacción atómica
                    with transaction.atomic():
                        # Verificar si el usuario ya existe
                        user_existente = User.objects.filter(username=documento_identidad).first()

                        if user_existente:
                            if actualizar:
                                # Actualizar usuario existente
                                user_existente.email = correo
                                user_existente.first_name = nombres
                                user_existente.last_name = f'{apellido1} {apellido2}'.strip()
                                user_existente.save()

                                # Actualizar persona
                                persona, created = Persona.objects.update_or_create(
                                    user=user_existente,
                                    defaults={
                                        'documento_identidad': documento_identidad,
                                        'nombres': nombres,
                                        'apellido1': apellido1,
                                        'apellido2': apellido2,
                                        'sexo': sexo,
                                        'fecha_nacimiento': fecha_nacimiento,
                                        'talla': talla,
                                        'numero_zapato': numero_zapato,
                                    }
                                )

                                if created:
                                    creadas += 1
                                else:
                                    actualizadas += 1
                            else:
                                self.stdout.write(
                                    self.style.WARNING(
                                        f'Fila {row_num}: Usuario {documento_identidad} ya existe, omitiendo...'
                                    )
                                )
                                continue
                        else:
                            # Crear nuevo usuario
                            user = User.objects.create(
                                username=documento_identidad,
                                email=correo,
                                first_name=nombres,
                                last_name=f'{apellido1} {apellido2}'.strip(),
                                password=PASSWORD_HASHEADA,
                                is_active=True,
                            )

                            # Crear persona vinculada
                            Persona.objects.create(
                                user=user,
                                documento_identidad=documento_identidad,
                                nombres=nombres,
                                apellido1=apellido1,
                                apellido2=apellido2,
                                sexo=sexo,
                                fecha_nacimiento=fecha_nacimiento,
                                talla=talla,
                                numero_zapato=numero_zapato,
                            )

                            creadas += 1

                    procesadas += 1

                    # Mostrar progreso cada 50 registros
                    if procesadas % 50 == 0:
                        self.stdout.write(f'Procesadas {procesadas}/{total_filas} filas...')

                except Exception as e:
                    errores += 1
                    self.stdout.write(
                        self.style.ERROR(f'Error en fila {row_num}: {str(e)}')
                    )
                    continue

            # Resumen final
            self.stdout.write(self.style.SUCCESS('\n=== Resumen de Importación ==='))
            self.stdout.write(f'Total procesadas: {procesadas}')
            self.stdout.write(f'Personas creadas: {creadas}')
            if actualizar:
                self.stdout.write(f'Personas actualizadas: {actualizadas}')
            self.stdout.write(f'Errores: {errores}')
            self.stdout.write(self.style.SUCCESS('\nImportación completada.'))

        except FileNotFoundError:
            self.stdout.write(
                self.style.ERROR(f'Archivo no encontrado: {archivo}')
            )
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f'Error durante la importación: {str(e)}')
            )
            import traceback
            self.stdout.write(traceback.format_exc())
