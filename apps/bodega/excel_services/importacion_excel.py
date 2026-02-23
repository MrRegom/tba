"""
Service Layer para importacion de datos desde Excel.

Contiene la logica de negocio para importar mantenedores desde archivos Excel.
Sigue Clean Architecture: separacion de responsabilidades.
"""
from typing import List, Dict, Any, Tuple
from openpyxl import load_workbook
from django.core.exceptions import ValidationError
from django.db import transaction


class ImportacionExcelService:
    """
    Service para importacion de datos desde Excel.
    
    Proporciona metodos genericos para:
    - Generar plantillas Excel
    - Validar archivos Excel
    - Importar datos desde Excel
    """
    
    @staticmethod
    def validar_archivo_excel(archivo) -> Tuple[bool, str]:
        """
        Valida que el archivo sea un Excel valido.
        
        Args:
            archivo: Archivo subido
            
        Returns:
            Tuple[bool, str]: (es_valido, mensaje_error)
        """
        if not archivo:
            return False, "No se proporciono archivo"
        
        if not archivo.name.endswith(('.xlsx', '.xls')):
            return False, "El archivo debe ser un Excel (.xlsx o .xls)"
        
        try:
            wb = load_workbook(archivo, read_only=True)
            wb.close()
            return True, ""
        except Exception as e:
            return False, f"Error al leer el archivo: {str(e)}"
    
    @staticmethod
    def leer_datos_desde_excel(archivo, columnas_esperadas: List[str], fila_inicio: int = 2) -> List[Dict[str, Any]]:
        """
        Lee datos desde un archivo Excel.
        
        Args:
            archivo: Archivo Excel subido
            columnas_esperadas: Lista de nombres de columnas esperadas
            fila_inicio: Fila donde comienzan los datos (default: 2, asumiendo fila 1 es encabezado)
            
        Returns:
            Lista de diccionarios con los datos leidos
        """
        wb = load_workbook(archivo, read_only=True)
        ws = wb.active
        
        datos = []
        
        # Leer encabezados de la primera fila
        encabezados = []
        for cell in ws[1]:
            encabezados.append(cell.value if cell.value else "")
        
        # Validar que las columnas esperadas esten presentes
        columnas_faltantes = [col for col in columnas_esperadas if col not in encabezados]
        if columnas_faltantes:
            wb.close()
            raise ValidationError(f"Columnas faltantes en el archivo: {', '.join(columnas_faltantes)}")
        
        # Leer datos desde la fila de inicio
        for row in ws.iter_rows(min_row=fila_inicio, values_only=False):
            # Saltar filas vacias
            if all(cell.value is None or str(cell.value).strip() == "" for cell in row):
                continue
            
            # Crear diccionario con los datos de la fila
            fila_data = {}
            for idx, header in enumerate(encabezados):
                if idx < len(row):
                    valor = row[idx].value
                    # Convertir None a string vacio
                    fila_data[header] = str(valor).strip() if valor is not None else ""
            
            datos.append(fila_data)
        
        wb.close()
        return datos
    
    @staticmethod
    def importar_marcas(archivo, usuario) -> Tuple[int, int, List[str]]:
        """
        Importa marcas desde un archivo Excel.
        
        Args:
            archivo: Archivo Excel con las marcas
            usuario: Usuario que realiza la importacion
            
        Returns:
            Tuple[int, int, List[str]]: (creadas, actualizadas, errores)
        """
        from apps.bodega.models import Marca
        
        columnas_esperadas = ['Codigo', 'Nombre', 'Descripcion', 'Activo']
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        
        creadas = 0
        actualizadas = 0
        errores = []
        
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):  # Empezar desde fila 2 (despues del encabezado)
                try:
                    codigo = fila.get('Codigo', '').strip()
                    nombre = fila.get('Nombre', '').strip()
                    descripcion = fila.get('Descripcion', '').strip()
                    activo_str = fila.get('Activo', 'SI').strip().upper()
                    
                    if not codigo or not nombre:
                        errores.append(f"Fila {idx}: Codigo y Nombre son obligatorios")
                        continue
                    
                    activo = activo_str in ['SI', 'S', 'TRUE', '1', 'ACTIVO']
                    
                    marca, created = Marca.objects.update_or_create(
                        codigo=codigo,
                        defaults={
                            'nombre': nombre,
                            'descripcion': descripcion,
                            'activo': activo,
                            'eliminado': False,
                        }
                    )
                    
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                        
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        
        return creadas, actualizadas, errores
    
    @staticmethod
    def generar_plantilla_marcas() -> bytes:
        """Genera plantilla de marcas con datos reales."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        from apps.bodega.models import Marca
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Marcas"
        
        encabezados = ['Codigo', 'Nombre', 'Descripcion', 'Activo']
        for col_idx, encabezado in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=encabezado)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Obtener datos existentes (hasta 10 registros)
        marcas_existentes = Marca.objects.filter(eliminado=False).order_by('codigo')[:10]
        
        for row_idx, marca in enumerate(marcas_existentes, start=2):
            ws.cell(row=row_idx, column=1, value=marca.codigo)
            ws.cell(row=row_idx, column=2, value=marca.nombre)
            ws.cell(row=row_idx, column=3, value=marca.descripcion)
            ws.cell(row=row_idx, column=4, value='SI' if marca.activo else 'NO')
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 10
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        
        return contenido
    
    @staticmethod
    def generar_plantilla_operaciones() -> bytes:
        """Genera plantilla de operaciones con datos reales."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        from apps.bodega.models import Operacion
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Operaciones"
        
        encabezados = ['Codigo', 'Nombre', 'Tipo', 'Descripcion', 'Activo']
        for col_idx, encabezado in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=encabezado)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Obtener datos existentes (hasta 10 registros)
        operaciones_existentes = Operacion.objects.filter(eliminado=False).order_by('codigo')[:10]
        
        for row_idx, operacion in enumerate(operaciones_existentes, start=2):
            ws.cell(row=row_idx, column=1, value=operacion.codigo)
            ws.cell(row=row_idx, column=2, value=operacion.nombre)
            ws.cell(row=row_idx, column=3, value=operacion.tipo)
            ws.cell(row=row_idx, column=4, value=operacion.descripcion)
            ws.cell(row=row_idx, column=5, value='SI' if operacion.activo else 'NO')
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 10
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        
        return contenido
    
    @staticmethod
    def importar_operaciones(archivo, usuario) -> Tuple[int, int, List[str]]:
        """Importa operaciones desde Excel."""
        from apps.bodega.models import Operacion
        
        columnas_esperadas = ['Codigo', 'Nombre', 'Tipo', 'Descripcion', 'Activo']
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        
        creadas = 0
        actualizadas = 0
        errores = []
        
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    codigo = fila.get('Codigo', '').strip()
                    nombre = fila.get('Nombre', '').strip()
                    tipo = fila.get('Tipo', '').strip()
                    descripcion = fila.get('Descripcion', '').strip()
                    activo_str = fila.get('Activo', 'SI').strip().upper()
                    
                    if not codigo or not nombre:
                        errores.append(f"Fila {idx}: Codigo y Nombre son obligatorios")
                        continue
                    
                    activo = activo_str in ['SI', 'S', 'TRUE', '1', 'ACTIVO']
                    
                    operacion, created = Operacion.objects.update_or_create(
                        codigo=codigo,
                        defaults={
                            'nombre': nombre,
                            'tipo': tipo,
                            'descripcion': descripcion,
                            'activo': activo,
                            'eliminado': False,
                        }
                    )
                    
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                        
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        
        return creadas, actualizadas, errores
    
    @staticmethod
    def generar_plantilla_tipos_movimiento() -> bytes:
        """Genera plantilla de tipos de movimiento con datos reales."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        from apps.bodega.models import TipoMovimiento
        
        wb = Workbook()
        ws = wb.active
        ws.title = "TiposMovimiento"
        
        encabezados = ['Codigo', 'Nombre', 'Descripcion', 'Activo']
        for col_idx, encabezado in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=encabezado)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Obtener datos existentes (hasta 10 registros)
        tipos_existentes = TipoMovimiento.objects.filter(eliminado=False).order_by('codigo')[:10]
        
        for row_idx, tipo in enumerate(tipos_existentes, start=2):
            ws.cell(row=row_idx, column=1, value=tipo.codigo)
            ws.cell(row=row_idx, column=2, value=tipo.nombre)
            ws.cell(row=row_idx, column=3, value=tipo.descripcion)
            ws.cell(row=row_idx, column=4, value='SI' if tipo.activo else 'NO')
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 10
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        
        return contenido
    
    @staticmethod
    def importar_tipos_movimiento(archivo, usuario) -> Tuple[int, int, List[str]]:
        """Importa tipos de movimiento desde Excel."""
        from apps.bodega.models import TipoMovimiento
        
        columnas_esperadas = ['Codigo', 'Nombre', 'Descripcion', 'Activo']
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        
        creadas = 0
        actualizadas = 0
        errores = []
        
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    codigo = fila.get('Codigo', '').strip()
                    nombre = fila.get('Nombre', '').strip()
                    descripcion = fila.get('Descripcion', '').strip()
                    activo_str = fila.get('Activo', 'SI').strip().upper()
                    
                    if not codigo or not nombre:
                        errores.append(f"Fila {idx}: Codigo y Nombre son obligatorios")
                        continue
                    
                    activo = activo_str in ['SI', 'S', 'TRUE', '1', 'ACTIVO']
                    
                    tipo, created = TipoMovimiento.objects.update_or_create(
                        codigo=codigo,
                        defaults={
                            'nombre': nombre,
                            'descripcion': descripcion,
                            'activo': activo,
                            'eliminado': False,
                        }
                    )
                    
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                        
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        
        return creadas, actualizadas, errores
    
    # ==================== METODOS PARA SOLICITUDES ====================
    
    @staticmethod
    def generar_plantilla_tipos_solicitud() -> bytes:
        """Genera plantilla de tipos de solicitud con datos reales."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        from apps.solicitudes.models import TipoSolicitud
        
        wb = Workbook()
        ws = wb.active
        ws.title = "TiposSolicitud"
        
        encabezados = ['Codigo', 'Nombre', 'Descripcion', 'RequiereAprobacion', 'Activo']
        for col_idx, encabezado in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=encabezado)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Obtener datos existentes (hasta 10 registros)
        tipos_existentes = TipoSolicitud.objects.filter(eliminado=False).order_by('codigo')[:10]
        
        for row_idx, tipo in enumerate(tipos_existentes, start=2):
            ws.cell(row=row_idx, column=1, value=tipo.codigo)
            ws.cell(row=row_idx, column=2, value=tipo.nombre)
            ws.cell(row=row_idx, column=3, value=tipo.descripcion or '')
            ws.cell(row=row_idx, column=4, value='SI' if tipo.requiere_aprobacion else 'NO')
            ws.cell(row=row_idx, column=5, value='SI' if tipo.activo else 'NO')
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 10
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        
        return contenido
    
    @staticmethod
    def importar_tipos_solicitud(archivo, usuario) -> Tuple[int, int, List[str]]:
        """Importa tipos de solicitud desde Excel."""
        from apps.solicitudes.models import TipoSolicitud
        
        columnas_esperadas = ['Codigo', 'Nombre', 'Descripcion', 'RequiereAprobacion', 'Activo']
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        
        creadas = 0
        actualizadas = 0
        errores = []
        
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    codigo = fila.get('Codigo', '').strip()
                    nombre = fila.get('Nombre', '').strip()
                    descripcion = fila.get('Descripcion', '').strip()
                    requiere_aprobacion_str = fila.get('RequiereAprobacion', 'SI').strip().upper()
                    activo_str = fila.get('Activo', 'SI').strip().upper()
                    
                    if not codigo or not nombre:
                        errores.append(f"Fila {idx}: Codigo y Nombre son obligatorios")
                        continue
                    
                    requiere_aprobacion = requiere_aprobacion_str in ['SI', 'S', 'TRUE', '1']
                    activo = activo_str in ['SI', 'S', 'TRUE', '1', 'ACTIVO']
                    
                    tipo, created = TipoSolicitud.objects.update_or_create(
                        codigo=codigo,
                        defaults={
                            'nombre': nombre,
                            'descripcion': descripcion,
                            'requiere_aprobacion': requiere_aprobacion,
                            'activo': activo,
                            'eliminado': False,
                        }
                    )
                    
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                        
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        
        return creadas, actualizadas, errores
    
    @staticmethod
    def generar_plantilla_estados_solicitud() -> bytes:
        """Genera plantilla de estados de solicitud con datos reales."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        from apps.solicitudes.models import EstadoSolicitud
        
        wb = Workbook()
        ws = wb.active
        ws.title = "EstadosSolicitud"
        
        encabezados = ['Codigo', 'Nombre', 'Descripcion', 'Color', 'Activo']
        for col_idx, encabezado in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=encabezado)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Obtener datos existentes (hasta 10 registros)
        estados_existentes = EstadoSolicitud.objects.filter(eliminado=False).order_by('codigo')[:10]
        
        for row_idx, estado in enumerate(estados_existentes, start=2):
            ws.cell(row=row_idx, column=1, value=estado.codigo)
            ws.cell(row=row_idx, column=2, value=estado.nombre)
            ws.cell(row=row_idx, column=3, value=estado.descripcion or '')
            ws.cell(row=row_idx, column=4, value=estado.color)
            ws.cell(row=row_idx, column=5, value='SI' if estado.activo else 'NO')
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        
        return contenido
    
    @staticmethod
    def importar_estados_solicitud(archivo, usuario) -> Tuple[int, int, List[str]]:
        """Importa estados de solicitud desde Excel."""
        from apps.solicitudes.models import EstadoSolicitud
        
        columnas_esperadas = ['Codigo', 'Nombre', 'Descripcion', 'Color', 'Activo']
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        
        creadas = 0
        actualizadas = 0
        errores = []
        
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    codigo = fila.get('Codigo', '').strip()
                    nombre = fila.get('Nombre', '').strip()
                    descripcion = fila.get('Descripcion', '').strip()
                    color = fila.get('Color', '#6c757d').strip()
                    activo_str = fila.get('Activo', 'SI').strip().upper()
                    
                    if not codigo or not nombre:
                        errores.append(f"Fila {idx}: Codigo y Nombre son obligatorios")
                        continue
                    
                    activo = activo_str in ['SI', 'S', 'TRUE', '1', 'ACTIVO']
                    
                    estado, created = EstadoSolicitud.objects.update_or_create(
                        codigo=codigo,
                        defaults={
                            'nombre': nombre,
                            'descripcion': descripcion,
                            'color': color,
                            'activo': activo,
                            'eliminado': False,
                        }
                    )
                    
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                        
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        
        return creadas, actualizadas, errores
    
    # ==================== METODOS PARA COMPRAS ====================
    
    @staticmethod
    def generar_plantilla_estados_recepcion() -> bytes:
        """Genera plantilla de estados de recepcion con datos reales."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        from apps.bodega.models import EstadoRecepcion
        
        wb = Workbook()
        ws = wb.active
        ws.title = "EstadosRecepcion"
        
        encabezados = ['Codigo', 'Nombre', 'Descripcion', 'Color', 'Activo']
        for col_idx, encabezado in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=encabezado)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Obtener datos existentes (hasta 10 registros)
        estados_existentes = EstadoRecepcion.objects.filter(eliminado=False).order_by('codigo')[:10]
        
        for row_idx, estado in enumerate(estados_existentes, start=2):
            ws.cell(row=row_idx, column=1, value=estado.codigo)
            ws.cell(row=row_idx, column=2, value=estado.nombre)
            ws.cell(row=row_idx, column=3, value=estado.descripcion or '')
            ws.cell(row=row_idx, column=4, value=estado.color)
            ws.cell(row=row_idx, column=5, value='SI' if estado.activo else 'NO')
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        
        return contenido
    
    @staticmethod
    def importar_estados_recepcion(archivo, usuario) -> Tuple[int, int, List[str]]:
        """Importa estados de recepcion desde Excel."""
        from apps.bodega.models import EstadoRecepcion
        
        columnas_esperadas = ['Codigo', 'Nombre', 'Descripcion', 'Color', 'Activo']
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        
        creadas = 0
        actualizadas = 0
        errores = []
        
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    codigo = fila.get('Codigo', '').strip()
                    nombre = fila.get('Nombre', '').strip()
                    descripcion = fila.get('Descripcion', '').strip()
                    color = fila.get('Color', '#6c757d').strip()
                    activo_str = fila.get('Activo', 'SI').strip().upper()
                    
                    if not codigo or not nombre:
                        errores.append(f"Fila {idx}: Codigo y Nombre son obligatorios")
                        continue
                    
                    activo = activo_str in ['SI', 'S', 'TRUE', '1', 'ACTIVO']
                    
                    estado, created = EstadoRecepcion.objects.update_or_create(
                        codigo=codigo,
                        defaults={
                            'nombre': nombre,
                            'descripcion': descripcion,
                            'color': color,
                            'activo': activo,
                            'eliminado': False,
                        }
                    )
                    
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                        
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        
        return creadas, actualizadas, errores
    
    @staticmethod
    def generar_plantilla_tipos_recepcion() -> bytes:
        """Genera plantilla de tipos de recepcion con datos reales."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        from apps.bodega.models import TipoRecepcion
        
        wb = Workbook()
        ws = wb.active
        ws.title = "TiposRecepcion"
        
        encabezados = ['Codigo', 'Nombre', 'Descripcion', 'RequiereOrden', 'Activo']
        for col_idx, encabezado in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=encabezado)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Obtener datos existentes (hasta 10 registros)
        tipos_existentes = TipoRecepcion.objects.filter(eliminado=False).order_by('codigo')[:10]
        
        for row_idx, tipo in enumerate(tipos_existentes, start=2):
            ws.cell(row=row_idx, column=1, value=tipo.codigo)
            ws.cell(row=row_idx, column=2, value=tipo.nombre)
            ws.cell(row=row_idx, column=3, value=tipo.descripcion or '')
            ws.cell(row=row_idx, column=4, value='SI' if tipo.requiere_orden else 'NO')
            ws.cell(row=row_idx, column=5, value='SI' if tipo.activo else 'NO')
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 10
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        
        return contenido
    
    @staticmethod
    def importar_tipos_recepcion(archivo, usuario) -> Tuple[int, int, List[str]]:
        """Importa tipos de recepcion desde Excel."""
        from apps.bodega.models import TipoRecepcion
        
        columnas_esperadas = ['Codigo', 'Nombre', 'Descripcion', 'RequiereOrden', 'Activo']
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        
        creadas = 0
        actualizadas = 0
        errores = []
        
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    codigo = fila.get('Codigo', '').strip()
                    nombre = fila.get('Nombre', '').strip()
                    descripcion = fila.get('Descripcion', '').strip()
                    requiere_orden_str = fila.get('RequiereOrden', 'NO').strip().upper()
                    activo_str = fila.get('Activo', 'SI').strip().upper()
                    
                    if not codigo or not nombre:
                        errores.append(f"Fila {idx}: Codigo y Nombre son obligatorios")
                        continue
                    
                    requiere_orden = requiere_orden_str in ['SI', 'S', 'TRUE', '1']
                    activo = activo_str in ['SI', 'S', 'TRUE', '1', 'ACTIVO']
                    
                    tipo, created = TipoRecepcion.objects.update_or_create(
                        codigo=codigo,
                        defaults={
                            'nombre': nombre,
                            'descripcion': descripcion,
                            'requiere_orden': requiere_orden,
                            'activo': activo,
                            'eliminado': False,
                        }
                    )
                    
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                        
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        
        return creadas, actualizadas, errores
    
    @staticmethod
    def generar_plantilla_estados_orden_compra() -> bytes:
        """Genera plantilla de estados de orden de compra con datos reales."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        from apps.compras.models import EstadoOrdenCompra
        
        wb = Workbook()
        ws = wb.active
        ws.title = "EstadosOrdenCompra"
        
        encabezados = ['Codigo', 'Nombre', 'Descripcion', 'Color', 'Activo']
        for col_idx, encabezado in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=encabezado)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Obtener datos existentes (hasta 10 registros)
        estados_existentes = EstadoOrdenCompra.objects.filter(eliminado=False).order_by('codigo')[:10]
        
        for row_idx, estado in enumerate(estados_existentes, start=2):
            ws.cell(row=row_idx, column=1, value=estado.codigo)
            ws.cell(row=row_idx, column=2, value=estado.nombre)
            ws.cell(row=row_idx, column=3, value=estado.descripcion or '')
            ws.cell(row=row_idx, column=4, value=estado.color)
            ws.cell(row=row_idx, column=5, value='SI' if estado.activo else 'NO')
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        
        return contenido
    
    @staticmethod
    def importar_estados_orden_compra(archivo, usuario) -> Tuple[int, int, List[str]]:
        """Importa estados de orden de compra desde Excel."""
        from apps.compras.models import EstadoOrdenCompra
        
        columnas_esperadas = ['Codigo', 'Nombre', 'Descripcion', 'Color', 'Activo']
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        
        creadas = 0
        actualizadas = 0
        errores = []
        
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    codigo = fila.get('Codigo', '').strip()
                    nombre = fila.get('Nombre', '').strip()
                    descripcion = fila.get('Descripcion', '').strip()
                    color = fila.get('Color', '#6c757d').strip()
                    activo_str = fila.get('Activo', 'SI').strip().upper()
                    
                    if not codigo or not nombre:
                        errores.append(f"Fila {idx}: Codigo y Nombre son obligatorios")
                        continue
                    
                    activo = activo_str in ['SI', 'S', 'TRUE', '1', 'ACTIVO']
                    
                    estado, created = EstadoOrdenCompra.objects.update_or_create(
                        codigo=codigo,
                        defaults={
                            'nombre': nombre,
                            'descripcion': descripcion,
                            'color': color,
                            'activo': activo,
                            'eliminado': False,
                        }
                    )
                    
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                        
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        
        return creadas, actualizadas, errores


    # ==================== MÉTODOS BODEGA (NUEVOS) ====================

    @staticmethod
    def generar_plantilla_categorias_bodega() -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        from apps.bodega.models import Categoria
        wb = Workbook()
        ws = wb.active
        ws.title = 'Categorias'
        encabezados = ['Codigo', 'Nombre', 'Descripcion', 'Activo']
        for col_idx, enc in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=enc)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for row_idx, obj in enumerate(Categoria.objects.filter(eliminado=False).order_by('codigo')[:10], start=2):
            ws.cell(row=row_idx, column=1, value=obj.codigo)
            ws.cell(row=row_idx, column=2, value=obj.nombre)
            ws.cell(row=row_idx, column=3, value=obj.descripcion or '')
            ws.cell(row=row_idx, column=4, value='SI' if obj.activo else 'NO')
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 10
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        return contenido

    @staticmethod
    def importar_categorias_bodega(archivo, usuario):
        from apps.bodega.models import Categoria
        columnas_esperadas = ['Codigo', 'Nombre', 'Descripcion', 'Activo']
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        creadas = 0
        actualizadas = 0
        errores = []
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    codigo = fila.get('Codigo', '').strip()
                    nombre = fila.get('Nombre', '').strip()
                    if not codigo or not nombre:
                        errores.append(f"Fila {idx}: Codigo y Nombre son obligatorios")
                        continue
                    activo = fila.get('Activo', 'SI').strip().upper() in ['SI', 'S', 'TRUE', '1', 'ACTIVO']
                    obj, created = Categoria.objects.update_or_create(codigo=codigo, defaults={'nombre': nombre, 'descripcion': fila.get('Descripcion', '').strip(), 'activo': activo, 'eliminado': False})
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        return creadas, actualizadas, errores

    @staticmethod
    def generar_plantilla_unidades_medida() -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        from apps.bodega.models import UnidadMedida
        wb = Workbook()
        ws = wb.active
        ws.title = 'UnidadesMedida'
        encabezados = ['Codigo', 'Nombre', 'Simbolo', 'Activo']
        for col_idx, enc in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=enc)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for row_idx, obj in enumerate(UnidadMedida.objects.filter(eliminado=False).order_by('codigo')[:10], start=2):
            ws.cell(row=row_idx, column=1, value=obj.codigo)
            ws.cell(row=row_idx, column=2, value=obj.nombre)
            ws.cell(row=row_idx, column=3, value=obj.simbolo)
            ws.cell(row=row_idx, column=4, value='SI' if obj.activo else 'NO')
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        return contenido

    @staticmethod
    def importar_unidades_medida(archivo, usuario):
        from apps.bodega.models import UnidadMedida
        columnas_esperadas = ['Codigo', 'Nombre', 'Simbolo', 'Activo']
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        creadas = 0
        actualizadas = 0
        errores = []
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    codigo = fila.get('Codigo', '').strip()
                    nombre = fila.get('Nombre', '').strip()
                    simbolo = fila.get('Simbolo', '').strip()
                    if not codigo or not nombre or not simbolo:
                        errores.append(f"Fila {idx}: Codigo, Nombre y Simbolo son obligatorios")
                        continue
                    activo = fila.get('Activo', 'SI').strip().upper() in ['SI', 'S', 'TRUE', '1', 'ACTIVO']
                    obj, created = UnidadMedida.objects.update_or_create(codigo=codigo, defaults={'nombre': nombre, 'simbolo': simbolo, 'activo': activo, 'eliminado': False})
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        return creadas, actualizadas, errores

    @staticmethod
    def generar_plantilla_articulos() -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        from apps.bodega.models import Articulo
        wb = Workbook()
        ws = wb.active
        ws.title = 'Articulos'
        encabezados = ['Codigo', 'Nombre', 'Descripcion', 'Categoria', 'Marca', 'UnidadMedida', 'StockMinimo', 'StockMaximo']
        for col_idx, enc in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=enc)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for row_idx, obj in enumerate(Articulo.objects.filter(eliminado=False).select_related('categoria', 'marca', 'unidad_medida').order_by('codigo')[:10], start=2):
            ws.cell(row=row_idx, column=1, value=obj.codigo)
            ws.cell(row=row_idx, column=2, value=obj.nombre)
            ws.cell(row=row_idx, column=3, value=obj.descripcion or '')
            ws.cell(row=row_idx, column=4, value=obj.categoria.codigo if obj.categoria else '')
            ws.cell(row=row_idx, column=5, value=obj.marca.codigo if obj.marca else '')
            ws.cell(row=row_idx, column=6, value=obj.unidad_medida.codigo if obj.unidad_medida else '')
            ws.cell(row=row_idx, column=7, value=obj.stock_minimo)
            ws.cell(row=row_idx, column=8, value=obj.stock_maximo or '')
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 20
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        return contenido

    @staticmethod
    def importar_articulos(archivo, usuario):
        from apps.bodega.models import Articulo, Categoria, Marca, UnidadMedida, Bodega
        columnas_esperadas = ['Codigo', 'Nombre', 'Descripcion', 'Categoria', 'Marca', 'UnidadMedida', 'StockMinimo', 'StockMaximo']
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        creadas = 0
        actualizadas = 0
        errores = []
        bodega_default = Bodega.objects.filter(eliminado=False).first()
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    codigo = fila.get('Codigo', '').strip()
                    nombre = fila.get('Nombre', '').strip()
                    if not codigo or not nombre:
                        errores.append(f"Fila {idx}: Codigo y Nombre son obligatorios")
                        continue
                    cat_codigo = fila.get('Categoria', '').strip()
                    if not cat_codigo:
                        errores.append(f"Fila {idx}: Categoria es obligatoria")
                        continue
                    try:
                        categoria = Categoria.objects.get(codigo=cat_codigo, eliminado=False)
                    except Categoria.DoesNotExist:
                        errores.append(f"Fila {idx}: Categoria '{cat_codigo}' no encontrada")
                        continue
                    marca = None
                    mk = fila.get('Marca', '').strip()
                    if mk:
                        marca = Marca.objects.filter(codigo=mk, eliminado=False).first()
                    unidad = None
                    um = fila.get('UnidadMedida', '').strip()
                    if um:
                        unidad = UnidadMedida.objects.filter(codigo=um, eliminado=False).first()
                    stock_min_str = fila.get('StockMinimo', '0').strip()
                    stock_min = int(stock_min_str) if stock_min_str and stock_min_str.isdigit() else 0
                    stock_max_str = fila.get('StockMaximo', '').strip()
                    stock_max = int(stock_max_str) if stock_max_str and stock_max_str.isdigit() else None
                    defaults = {'nombre': nombre, 'descripcion': fila.get('Descripcion', '').strip(), 'categoria': categoria, 'marca': marca, 'unidad_medida': unidad, 'stock_minimo': stock_min, 'stock_maximo': stock_max, 'eliminado': False}
                    if bodega_default:
                        defaults['ubicacion_fisica'] = bodega_default
                    obj, created = Articulo.objects.update_or_create(codigo=codigo, defaults=defaults)
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        return creadas, actualizadas, errores

    # ==================== MÉTODOS PARA ORGANIZACIÓN ====================

    @staticmethod
    def generar_plantilla_ubicaciones() -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        from apps.activos.models import Ubicacion
        wb = Workbook()
        ws = wb.active
        ws.title = 'Ubicaciones'
        encabezados = ['Codigo', 'Nombre', 'Descripcion', 'Activo']
        for col_idx, enc in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=enc)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for row_idx, obj in enumerate(Ubicacion.objects.filter(eliminado=False).order_by('codigo')[:10], start=2):
            ws.cell(row=row_idx, column=1, value=obj.codigo)
            ws.cell(row=row_idx, column=2, value=obj.nombre)
            ws.cell(row=row_idx, column=3, value=obj.descripcion or '')
            ws.cell(row=row_idx, column=4, value='SI' if obj.activo else 'NO')
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 10
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        return contenido

    @staticmethod
    def importar_ubicaciones(archivo, usuario):
        from apps.activos.models import Ubicacion
        columnas_esperadas = ['Codigo', 'Nombre', 'Descripcion', 'Activo']
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        creadas = 0
        actualizadas = 0
        errores = []
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    codigo = fila.get('Codigo', '').strip()
                    nombre = fila.get('Nombre', '').strip()
                    if not codigo or not nombre:
                        errores.append(f"Fila {idx}: Codigo y Nombre son obligatorios")
                        continue
                    activo = fila.get('Activo', 'SI').strip().upper() in ['SI', 'S', 'TRUE', '1', 'ACTIVO']
                    obj, created = Ubicacion.objects.update_or_create(codigo=codigo, defaults={'nombre': nombre, 'descripcion': fila.get('Descripcion', '').strip(), 'activo': activo, 'eliminado': False})
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        return creadas, actualizadas, errores

    @staticmethod
    def generar_plantilla_talleres() -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        from apps.activos.models import Taller
        wb = Workbook()
        ws = wb.active
        ws.title = 'Talleres'
        encabezados = ['Codigo', 'Nombre', 'Descripcion', 'Ubicacion', 'Responsable', 'Activo']
        for col_idx, enc in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=enc)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for row_idx, obj in enumerate(Taller.objects.filter(eliminado=False).select_related('responsable').order_by('codigo')[:10], start=2):
            ws.cell(row=row_idx, column=1, value=obj.codigo)
            ws.cell(row=row_idx, column=2, value=obj.nombre)
            ws.cell(row=row_idx, column=3, value=obj.descripcion or '')
            ws.cell(row=row_idx, column=4, value=obj.ubicacion or '')
            ws.cell(row=row_idx, column=5, value=obj.responsable.username if obj.responsable else '')
            ws.cell(row=row_idx, column=6, value='SI' if obj.activo else 'NO')
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 25
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        return contenido

    @staticmethod
    def importar_talleres(archivo, usuario):
        from django.contrib.auth import get_user_model
        from apps.activos.models import Taller
        User = get_user_model()
        columnas_esperadas = ['Codigo', 'Nombre', 'Descripcion', 'Ubicacion', 'Responsable', 'Activo']
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        creadas = 0
        actualizadas = 0
        errores = []
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    codigo = fila.get('Codigo', '').strip()
                    nombre = fila.get('Nombre', '').strip()
                    if not codigo or not nombre:
                        errores.append(f"Fila {idx}: Codigo y Nombre son obligatorios")
                        continue
                    responsable = None
                    resp = fila.get('Responsable', '').strip()
                    if resp:
                        responsable = User.objects.filter(username=resp).first()
                    activo = fila.get('Activo', 'SI').strip().upper() in ['SI', 'S', 'TRUE', '1', 'ACTIVO']
                    obj, created = Taller.objects.update_or_create(codigo=codigo, defaults={'nombre': nombre, 'descripcion': fila.get('Descripcion', '').strip(), 'ubicacion': fila.get('Ubicacion', '').strip(), 'responsable': responsable, 'activo': activo, 'eliminado': False})
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        return creadas, actualizadas, errores

    @staticmethod
    def generar_plantilla_bodegas() -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        from apps.bodega.models import Bodega
        wb = Workbook()
        ws = wb.active
        ws.title = 'Bodegas'
        encabezados = ['Codigo', 'Nombre', 'Descripcion', 'Responsable']
        for col_idx, enc in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=enc)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for row_idx, obj in enumerate(Bodega.objects.filter(eliminado=False).select_related('responsable').order_by('codigo')[:10], start=2):
            ws.cell(row=row_idx, column=1, value=obj.codigo)
            ws.cell(row=row_idx, column=2, value=obj.nombre)
            ws.cell(row=row_idx, column=3, value=obj.descripcion or '')
            ws.cell(row=row_idx, column=4, value=obj.responsable.username if obj.responsable else '')
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 25
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        return contenido

    @staticmethod
    def importar_bodegas(archivo, usuario):
        from django.contrib.auth import get_user_model
        from apps.bodega.models import Bodega
        User = get_user_model()
        columnas_esperadas = ['Codigo', 'Nombre', 'Descripcion', 'Responsable']
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        creadas = 0
        actualizadas = 0
        errores = []
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    codigo = fila.get('Codigo', '').strip()
                    nombre = fila.get('Nombre', '').strip()
                    if not codigo or not nombre:
                        errores.append(f"Fila {idx}: Codigo y Nombre son obligatorios")
                        continue
                    responsable = None
                    resp = fila.get('Responsable', '').strip()
                    if resp:
                        responsable = User.objects.filter(username=resp).first()
                    if not responsable:
                        responsable = usuario
                    obj, created = Bodega.objects.update_or_create(codigo=codigo, defaults={'nombre': nombre, 'descripcion': fila.get('Descripcion', '').strip(), 'responsable': responsable, 'eliminado': False})
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        return creadas, actualizadas, errores

    @staticmethod
    def generar_plantilla_departamentos() -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        from apps.solicitudes.models import Departamento
        wb = Workbook()
        ws = wb.active
        ws.title = 'Departamentos'
        encabezados = ['Codigo', 'Nombre', 'Descripcion', 'Responsable', 'Activo']
        for col_idx, enc in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=enc)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for row_idx, obj in enumerate(Departamento.objects.filter(eliminado=False).select_related('responsable').order_by('codigo')[:10], start=2):
            ws.cell(row=row_idx, column=1, value=obj.codigo)
            ws.cell(row=row_idx, column=2, value=obj.nombre)
            ws.cell(row=row_idx, column=3, value=obj.descripcion or '')
            ws.cell(row=row_idx, column=4, value=obj.responsable.username if obj.responsable else '')
            ws.cell(row=row_idx, column=5, value='SI' if obj.activo else 'NO')
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 25
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        return contenido

    @staticmethod
    def importar_departamentos(archivo, usuario):
        from django.contrib.auth import get_user_model
        from apps.solicitudes.models import Departamento
        User = get_user_model()
        columnas_esperadas = ['Codigo', 'Nombre', 'Descripcion', 'Responsable', 'Activo']
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        creadas = 0
        actualizadas = 0
        errores = []
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    codigo = fila.get('Codigo', '').strip()
                    nombre = fila.get('Nombre', '').strip()
                    if not codigo or not nombre:
                        errores.append(f"Fila {idx}: Codigo y Nombre son obligatorios")
                        continue
                    responsable = None
                    resp = fila.get('Responsable', '').strip()
                    if resp:
                        responsable = User.objects.filter(username=resp).first()
                    activo = fila.get('Activo', 'SI').strip().upper() in ['SI', 'S', 'TRUE', '1', 'ACTIVO']
                    obj, created = Departamento.objects.update_or_create(codigo=codigo, defaults={'nombre': nombre, 'descripcion': fila.get('Descripcion', '').strip(), 'responsable': responsable, 'activo': activo, 'eliminado': False})
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        return creadas, actualizadas, errores

    @staticmethod
    def generar_plantilla_areas() -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        from apps.solicitudes.models import Area
        wb = Workbook()
        ws = wb.active
        ws.title = 'Areas'
        encabezados = ['Codigo', 'Nombre', 'Descripcion', 'Departamento', 'Responsable', 'Activo']
        for col_idx, enc in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=enc)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for row_idx, obj in enumerate(Area.objects.filter(eliminado=False).select_related('departamento', 'responsable').order_by('codigo')[:10], start=2):
            ws.cell(row=row_idx, column=1, value=obj.codigo)
            ws.cell(row=row_idx, column=2, value=obj.nombre)
            ws.cell(row=row_idx, column=3, value=obj.descripcion or '')
            ws.cell(row=row_idx, column=4, value=obj.departamento.codigo if obj.departamento else '')
            ws.cell(row=row_idx, column=5, value=obj.responsable.username if obj.responsable else '')
            ws.cell(row=row_idx, column=6, value='SI' if obj.activo else 'NO')
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 25
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        return contenido

    @staticmethod
    def importar_areas(archivo, usuario):
        from django.contrib.auth import get_user_model
        from apps.solicitudes.models import Area, Departamento
        User = get_user_model()
        columnas_esperadas = ['Codigo', 'Nombre', 'Descripcion', 'Departamento', 'Responsable', 'Activo']
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        creadas = 0
        actualizadas = 0
        errores = []
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    codigo = fila.get('Codigo', '').strip()
                    nombre = fila.get('Nombre', '').strip()
                    depto_codigo = fila.get('Departamento', '').strip()
                    if not codigo or not nombre:
                        errores.append(f"Fila {idx}: Codigo y Nombre son obligatorios")
                        continue
                    if not depto_codigo:
                        errores.append(f"Fila {idx}: Departamento es obligatorio")
                        continue
                    try:
                        departamento = Departamento.objects.get(codigo=depto_codigo, eliminado=False)
                    except Departamento.DoesNotExist:
                        errores.append(f"Fila {idx}: Departamento '{depto_codigo}' no encontrado")
                        continue
                    responsable = None
                    resp = fila.get('Responsable', '').strip()
                    if resp:
                        responsable = User.objects.filter(username=resp).first()
                    activo = fila.get('Activo', 'SI').strip().upper() in ['SI', 'S', 'TRUE', '1', 'ACTIVO']
                    obj, created = Area.objects.update_or_create(codigo=codigo, defaults={'nombre': nombre, 'descripcion': fila.get('Descripcion', '').strip(), 'departamento': departamento, 'responsable': responsable, 'activo': activo, 'eliminado': False})
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        return creadas, actualizadas, errores

    # ==================== MÉTODOS PARA INVENTARIO (ACTIVOS) ====================

    @staticmethod
    def generar_plantilla_categorias_activo() -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        from apps.activos.models import CategoriaActivo
        wb = Workbook()
        ws = wb.active
        ws.title = 'CategoriasActivo'
        encabezados = ['Codigo', 'Nombre', 'Sigla', 'Descripcion', 'Activo']
        for col_idx, enc in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=enc)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for row_idx, obj in enumerate(CategoriaActivo.objects.filter(eliminado=False).order_by('codigo')[:10], start=2):
            ws.cell(row=row_idx, column=1, value=obj.codigo)
            ws.cell(row=row_idx, column=2, value=obj.nombre)
            ws.cell(row=row_idx, column=3, value=obj.sigla)
            ws.cell(row=row_idx, column=4, value=obj.descripcion or '')
            ws.cell(row=row_idx, column=5, value='SI' if obj.activo else 'NO')
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 10
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        return contenido

    @staticmethod
    def importar_categorias_activo(archivo, usuario):
        from apps.activos.models import CategoriaActivo
        columnas_esperadas = ['Codigo', 'Nombre', 'Sigla', 'Descripcion', 'Activo']
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        creadas = 0
        actualizadas = 0
        errores = []
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    codigo = fila.get('Codigo', '').strip()
                    nombre = fila.get('Nombre', '').strip()
                    sigla = fila.get('Sigla', '').strip()
                    if not codigo or not nombre or not sigla:
                        errores.append(f"Fila {idx}: Codigo, Nombre y Sigla son obligatorios")
                        continue
                    activo = fila.get('Activo', 'SI').strip().upper() in ['SI', 'S', 'TRUE', '1', 'ACTIVO']
                    obj, created = CategoriaActivo.objects.update_or_create(codigo=codigo, defaults={'nombre': nombre, 'sigla': sigla, 'descripcion': fila.get('Descripcion', '').strip(), 'activo': activo, 'eliminado': False})
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        return creadas, actualizadas, errores

    @staticmethod
    def generar_plantilla_estados_activo() -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        from apps.activos.models import EstadoActivo
        wb = Workbook()
        ws = wb.active
        ws.title = 'EstadosActivo'
        encabezados = ['Codigo', 'Nombre', 'Descripcion', 'Color', 'EsInicial', 'PermiteMovimiento', 'Activo']
        for col_idx, enc in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=enc)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for row_idx, obj in enumerate(EstadoActivo.objects.filter(eliminado=False).order_by('codigo')[:10], start=2):
            ws.cell(row=row_idx, column=1, value=obj.codigo)
            ws.cell(row=row_idx, column=2, value=obj.nombre)
            ws.cell(row=row_idx, column=3, value=obj.descripcion or '')
            ws.cell(row=row_idx, column=4, value=obj.color)
            ws.cell(row=row_idx, column=5, value='SI' if obj.es_inicial else 'NO')
            ws.cell(row=row_idx, column=6, value='SI' if obj.permite_movimiento else 'NO')
            ws.cell(row=row_idx, column=7, value='SI' if obj.activo else 'NO')
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 20
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        return contenido

    @staticmethod
    def importar_estados_activo(archivo, usuario):
        from apps.activos.models import EstadoActivo
        columnas_esperadas = ['Codigo', 'Nombre', 'Descripcion', 'Color', 'EsInicial', 'PermiteMovimiento', 'Activo']
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        creadas = 0
        actualizadas = 0
        errores = []
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    codigo = fila.get('Codigo', '').strip()
                    nombre = fila.get('Nombre', '').strip()
                    if not codigo or not nombre:
                        errores.append(f"Fila {idx}: Codigo y Nombre son obligatorios")
                        continue
                    color = fila.get('Color', '#6c757d').strip() or '#6c757d'
                    es_inicial = fila.get('EsInicial', 'NO').strip().upper() in ['SI', 'S', 'TRUE', '1']
                    permite_mov = fila.get('PermiteMovimiento', 'SI').strip().upper() in ['SI', 'S', 'TRUE', '1']
                    activo = fila.get('Activo', 'SI').strip().upper() in ['SI', 'S', 'TRUE', '1', 'ACTIVO']
                    obj, created = EstadoActivo.objects.update_or_create(codigo=codigo, defaults={'nombre': nombre, 'descripcion': fila.get('Descripcion', '').strip(), 'color': color, 'es_inicial': es_inicial, 'permite_movimiento': permite_mov, 'activo': activo, 'eliminado': False})
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        return creadas, actualizadas, errores

    @staticmethod
    def generar_plantilla_marcas_activo() -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        from apps.activos.models import Marca
        wb = Workbook()
        ws = wb.active
        ws.title = 'MarcasActivo'
        encabezados = ['Codigo', 'Nombre', 'Descripcion', 'Activo']
        for col_idx, enc in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=enc)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for row_idx, obj in enumerate(Marca.objects.filter(eliminado=False).order_by('codigo')[:10], start=2):
            ws.cell(row=row_idx, column=1, value=obj.codigo)
            ws.cell(row=row_idx, column=2, value=obj.nombre)
            ws.cell(row=row_idx, column=3, value=obj.descripcion or '')
            ws.cell(row=row_idx, column=4, value='SI' if obj.activo else 'NO')
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 10
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        return contenido

    @staticmethod
    def importar_marcas_activo(archivo, usuario):
        from apps.activos.models import Marca
        columnas_esperadas = ['Codigo', 'Nombre', 'Descripcion', 'Activo']
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        creadas = 0
        actualizadas = 0
        errores = []
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    codigo = fila.get('Codigo', '').strip()
                    nombre = fila.get('Nombre', '').strip()
                    if not codigo or not nombre:
                        errores.append(f"Fila {idx}: Codigo y Nombre son obligatorios")
                        continue
                    activo = fila.get('Activo', 'SI').strip().upper() in ['SI', 'S', 'TRUE', '1', 'ACTIVO']
                    obj, created = Marca.objects.update_or_create(codigo=codigo, defaults={'nombre': nombre, 'descripcion': fila.get('Descripcion', '').strip(), 'activo': activo, 'eliminado': False})
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        return creadas, actualizadas, errores

    @staticmethod
    def generar_plantilla_tipos_movimiento_activo() -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        from apps.activos.models import TipoMovimientoActivo
        wb = Workbook()
        ws = wb.active
        ws.title = 'TiposMovActivo'
        encabezados = ['Codigo', 'Nombre', 'Descripcion', 'Activo']
        for col_idx, enc in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=enc)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for row_idx, obj in enumerate(TipoMovimientoActivo.objects.filter(eliminado=False).order_by('codigo')[:10], start=2):
            ws.cell(row=row_idx, column=1, value=obj.codigo)
            ws.cell(row=row_idx, column=2, value=obj.nombre)
            ws.cell(row=row_idx, column=3, value=obj.descripcion or '')
            ws.cell(row=row_idx, column=4, value='SI' if obj.activo else 'NO')
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 10
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        return contenido

    @staticmethod
    def importar_tipos_movimiento_activo(archivo, usuario):
        from apps.activos.models import TipoMovimientoActivo
        columnas_esperadas = ['Codigo', 'Nombre', 'Descripcion', 'Activo']
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        creadas = 0
        actualizadas = 0
        errores = []
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    codigo = fila.get('Codigo', '').strip()
                    nombre = fila.get('Nombre', '').strip()
                    if not codigo or not nombre:
                        errores.append(f"Fila {idx}: Codigo y Nombre son obligatorios")
                        continue
                    activo = fila.get('Activo', 'SI').strip().upper() in ['SI', 'S', 'TRUE', '1', 'ACTIVO']
                    obj, created = TipoMovimientoActivo.objects.update_or_create(codigo=codigo, defaults={'nombre': nombre, 'descripcion': fila.get('Descripcion', '').strip(), 'activo': activo, 'eliminado': False})
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        return creadas, actualizadas, errores

    @staticmethod
    def generar_plantilla_activos() -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        from apps.activos.models import Activo
        wb = Workbook()
        ws = wb.active
        ws.title = 'Activos'
        encabezados = ['Codigo', 'Nombre', 'Descripcion', 'Categoria', 'Estado', 'Marca', 'NumeroSerie']
        for col_idx, enc in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=enc)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for row_idx, obj in enumerate(Activo.objects.filter(eliminado=False).select_related('categoria', 'estado', 'marca').order_by('codigo')[:10], start=2):
            ws.cell(row=row_idx, column=1, value=obj.codigo)
            ws.cell(row=row_idx, column=2, value=obj.nombre)
            ws.cell(row=row_idx, column=3, value=obj.descripcion or '')
            ws.cell(row=row_idx, column=4, value=obj.categoria.codigo if obj.categoria else '')
            ws.cell(row=row_idx, column=5, value=obj.estado.codigo if obj.estado else '')
            ws.cell(row=row_idx, column=6, value=obj.marca.codigo if obj.marca else '')
            ws.cell(row=row_idx, column=7, value=obj.numero_serie or '')
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 22
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        return contenido

    @staticmethod
    def importar_activos(archivo, usuario):
        from apps.activos.models import Activo, CategoriaActivo, EstadoActivo, Marca
        columnas_esperadas = ['Codigo', 'Nombre', 'Descripcion', 'Categoria', 'Estado', 'Marca', 'NumeroSerie']
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        creadas = 0
        actualizadas = 0
        errores = []
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    codigo = fila.get('Codigo', '').strip()
                    nombre = fila.get('Nombre', '').strip()
                    if not codigo or not nombre:
                        errores.append(f"Fila {idx}: Codigo y Nombre son obligatorios")
                        continue
                    cat_codigo = fila.get('Categoria', '').strip()
                    if not cat_codigo:
                        errores.append(f"Fila {idx}: Categoria es obligatoria")
                        continue
                    try:
                        categoria = CategoriaActivo.objects.get(codigo=cat_codigo, eliminado=False)
                    except CategoriaActivo.DoesNotExist:
                        errores.append(f"Fila {idx}: Categoria '{cat_codigo}' no encontrada")
                        continue
                    est_codigo = fila.get('Estado', '').strip()
                    if not est_codigo:
                        errores.append(f"Fila {idx}: Estado es obligatorio")
                        continue
                    try:
                        estado = EstadoActivo.objects.get(codigo=est_codigo, eliminado=False)
                    except EstadoActivo.DoesNotExist:
                        errores.append(f"Fila {idx}: Estado '{est_codigo}' no encontrado")
                        continue
                    marca = None
                    mk = fila.get('Marca', '').strip()
                    if mk:
                        marca = Marca.objects.filter(codigo=mk, eliminado=False).first()
                    obj, created = Activo.objects.update_or_create(codigo=codigo, defaults={'nombre': nombre, 'descripcion': fila.get('Descripcion', '').strip(), 'categoria': categoria, 'estado': estado, 'marca': marca, 'numero_serie': fila.get('NumeroSerie', '').strip() or None, 'eliminado': False})
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        return creadas, actualizadas, errores

    # ==================== MÉTODOS PARA COMPRAS (PROVEEDOR) ====================

    @staticmethod
    def generar_plantilla_proveedores() -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        from apps.compras.models import Proveedor
        wb = Workbook()
        ws = wb.active
        ws.title = 'Proveedores'
        encabezados = ['RUT', 'RazonSocial', 'Direccion', 'Comuna', 'Ciudad', 'Telefono', 'Email', 'SitioWeb', 'Activo']
        for col_idx, enc in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=enc)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for row_idx, obj in enumerate(Proveedor.objects.filter(eliminado=False).order_by('rut')[:10], start=2):
            ws.cell(row=row_idx, column=1, value=obj.rut)
            ws.cell(row=row_idx, column=2, value=obj.razon_social)
            ws.cell(row=row_idx, column=3, value=obj.direccion)
            ws.cell(row=row_idx, column=4, value=obj.comuna or '')
            ws.cell(row=row_idx, column=5, value=obj.ciudad or '')
            ws.cell(row=row_idx, column=6, value=obj.telefono or '')
            ws.cell(row=row_idx, column=7, value=obj.email or '')
            ws.cell(row=row_idx, column=8, value=obj.sitio_web or '')
            ws.cell(row=row_idx, column=9, value='SI' if obj.activo else 'NO')
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws.column_dimensions[col].width = 22
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        return contenido

    @staticmethod
    def importar_proveedores(archivo, usuario):
        from apps.compras.models import Proveedor
        columnas_esperadas = ['RUT', 'RazonSocial', 'Direccion', 'Comuna', 'Ciudad', 'Telefono', 'Email', 'SitioWeb', 'Activo']
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        creadas = 0
        actualizadas = 0
        errores = []
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    rut = fila.get('RUT', '').strip()
                    razon_social = fila.get('RazonSocial', '').strip()
                    if not rut or not razon_social:
                        errores.append(f"Fila {idx}: RUT y RazonSocial son obligatorios")
                        continue
                    activo = fila.get('Activo', 'SI').strip().upper() in ['SI', 'S', 'TRUE', '1', 'ACTIVO']
                    obj, created = Proveedor.objects.update_or_create(rut=rut, defaults={'razon_social': razon_social, 'direccion': fila.get('Direccion', '').strip() or '-', 'comuna': fila.get('Comuna', '').strip() or None, 'ciudad': fila.get('Ciudad', '').strip() or None, 'telefono': fila.get('Telefono', '').strip() or None, 'email': fila.get('Email', '').strip() or None, 'sitio_web': fila.get('SitioWeb', '').strip() or None, 'activo': activo, 'eliminado': False})
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        return creadas, actualizadas, errores

    # ==================== MÉTODOS PARA USUARIOS (CARGO) ====================

    @staticmethod
    def generar_plantilla_cargos() -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        from apps.accounts.models import Cargo
        wb = Workbook()
        ws = wb.active
        ws.title = 'Cargos'
        encabezados = ['Codigo', 'Nombre', 'Activo']
        for col_idx, enc in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=enc)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for row_idx, obj in enumerate(Cargo.objects.filter(eliminado=False).order_by('codigo')[:10], start=2):
            ws.cell(row=row_idx, column=1, value=obj.codigo)
            ws.cell(row=row_idx, column=2, value=obj.nombre)
            ws.cell(row=row_idx, column=3, value='SI' if obj.activo else 'NO')
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 10
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        return contenido

    @staticmethod
    def importar_cargos(archivo, usuario):
        from apps.accounts.models import Cargo
        columnas_esperadas = ['Codigo', 'Nombre', 'Activo']
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        creadas = 0
        actualizadas = 0
        errores = []
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    codigo = fila.get('Codigo', '').strip()
                    nombre = fila.get('Nombre', '').strip()
                    if not codigo or not nombre:
                        errores.append(f"Fila {idx}: Codigo y Nombre son obligatorios")
                        continue
                    activo = fila.get('Activo', 'SI').strip().upper() in ['SI', 'S', 'TRUE', '1', 'ACTIVO']
                    obj, created = Cargo.objects.update_or_create(codigo=codigo, defaults={'nombre': nombre, 'activo': activo, 'eliminado': False})
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        return creadas, actualizadas, errores
