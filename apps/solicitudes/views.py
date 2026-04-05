"""
Class-Based Views para el módulo de solicitudes.

Este archivo implementa todas las vistas usando CBVs siguiendo SOLID y DRY:
- Reutilización de mixins de core.mixins
- Separación de responsabilidades (Repository Pattern + Service Layer)
- Código limpio y mantenible
- Type hints completos
- Auditoría automática
- Workflow de aprobación y despacho
"""
from typing import Any
from django.db.models import QuerySet, Q
from django.urls import reverse_lazy
from django.shortcuts import redirect, render
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.views.generic import (
    TemplateView, ListView, DetailView, CreateView, UpdateView, DeleteView, View
)
from core.mixins import (
    BaseAuditedViewMixin, AtomicTransactionMixin, SoftDeleteMixin,
    ScopedObjectPermissionMixin,
    PaginatedListMixin, FilteredListMixin
)
from core.utils import registrar_log_auditoria
from core.authz import can_view_solicitud, scope_solicitudes_for_user
from .mixins import (
    GestionSolicitudesPermissionMixin,
    AprobarSolicitudesPermissionMixin,
    RechazarSolicitudesPermissionMixin,
    DespacharSolicitudesPermissionMixin,
    CrearSolicitudArticulosPermissionMixin,
    CrearSolicitudBienesPermissionMixin,
    VerSolicitudesArticulosPermissionMixin,
    VerSolicitudesBienesPermissionMixin,
    MisSolicitudesPermissionMixin,
    EditarMisSolicitudesPermissionMixin,
    EliminarMisSolicitudesPermissionMixin,
)
from .models import Solicitud, TipoSolicitud, EstadoSolicitud, DetalleSolicitud, HistorialSolicitud
from .forms import (
    SolicitudForm, DetalleSolicitudArticuloFormSet, DetalleSolicitudActivoFormSet,
    AprobarSolicitudForm, DespacharSolicitudForm, RechazarSolicitudForm,
    FiltroSolicitudesForm, TipoSolicitudForm, EstadoSolicitudForm
)
from .repositories import (
    TipoSolicitudRepository, EstadoSolicitudRepository, SolicitudRepository,
    DetalleSolicitudRepository, HistorialSolicitudRepository
)
from .services import SolicitudService, DetalleSolicitudService
from decimal import Decimal


# ==================== VISTA MENÚ PRINCIPAL ====================

class MenuSolicitudesView(BaseAuditedViewMixin, TemplateView):
    """
    Vista del menú principal del módulo de solicitudes.

    Muestra estadísticas y accesos rápidos basados en permisos del usuario.
    Permisos: solicitudes.view_solicitud
    Utiliza: Repositories para acceso a datos optimizado
    """
    template_name = 'solicitudes/menu_solicitudes.html'
    permission_required = 'solicitudes.view_solicitud'

    def get_context_data(self, **kwargs) -> dict:
        """Agrega estadísticas y permisos al contexto usando repositories."""
        context = super().get_context_data(**kwargs)
        user = self.request.user

        # Inicializar repositories
        solicitud_repo = SolicitudRepository()
        estado_repo = EstadoSolicitudRepository()
        tipo_repo = TipoSolicitudRepository()

        # Estadísticas del módulo usando repositories
        estado_pendiente = estado_repo.get_by_codigo('PENDIENTE')
        pendientes_count = 0
        if estado_pendiente:
            pendientes_count = solicitud_repo.filter_by_estado(estado_pendiente).count()

        context['stats'] = {
            'total_solicitudes': solicitud_repo.get_all().count(),
            'mis_solicitudes': solicitud_repo.filter_by_solicitante(user).count(),
            'solicitudes_activos': solicitud_repo.filter_by_tipo_choice('ACTIVO').count(),
            'solicitudes_articulos': solicitud_repo.filter_by_tipo_choice('ARTICULO').count(),
            'pendientes': pendientes_count,
            # Estadísticas de mantenedores
            'total_tipos_solicitud': tipo_repo.get_all().count(),
            'total_estados_solicitud': estado_repo.get_all().count(),
        }

        # Permisos del usuario (usando nuevos permisos personalizados)
        context['permisos'] = {
            # Permisos de gestión
            'puede_gestionar': user.has_perm('solicitudes.gestionar_solicitudes'),
            'puede_aprobar': user.has_perm('solicitudes.aprobar_solicitudes'),
            'puede_rechazar': user.has_perm('solicitudes.rechazar_solicitudes'),
            'puede_despachar': user.has_perm('solicitudes.despachar_solicitudes'),
            'puede_ver_todas': user.has_perm('solicitudes.ver_todas_solicitudes'),

            # Permisos de solicitud de artículos
            'puede_crear_articulos': user.has_perm('solicitudes.crear_solicitud_articulos'),
            'puede_ver_solicitudes_articulos': user.has_perm('solicitudes.ver_solicitudes_articulos'),

            # Permisos de solicitud de bienes
            'puede_crear_bienes': user.has_perm('solicitudes.crear_solicitud_bienes'),
            'puede_ver_solicitudes_bienes': user.has_perm('solicitudes.ver_solicitudes_bienes'),

            # Permisos de mis solicitudes
            'puede_ver_mis_solicitudes': user.has_perm('solicitudes.ver_mis_solicitudes'),
            'puede_editar_mis_solicitudes': user.has_perm('solicitudes.editar_mis_solicitudes'),
            'puede_eliminar_mis_solicitudes': user.has_perm('solicitudes.eliminar_mis_solicitudes'),

            # Permisos para mantenedores
            'puede_gestionar_mantenedores': (
                user.has_perm('solicitudes.view_tiposolicitud') or
                user.has_perm('solicitudes.change_tiposolicitud') or
                user.has_perm('solicitudes.view_estadosolicitud') or
                user.has_perm('solicitudes.change_estadosolicitud')
            ),
        }

        context['titulo'] = 'Gestores - Solicitudes'

        # Datos para tabs (Tipos y Estados inline)
        context['tipos_solicitud'] = TipoSolicitud.objects.filter(eliminado=False).order_by('codigo')
        context['estados_solicitud'] = EstadoSolicitud.objects.filter(eliminado=False).order_by('codigo')

        return context


# ==================== VISTAS DE SOLICITUDES GENERALES ====================

class SolicitudListView(GestionSolicitudesPermissionMixin, BaseAuditedViewMixin, PaginatedListMixin, ListView):
    """
    Vista para listar todas las solicitudes con filtros (GESTIÓN).

    Permisos: solicitudes.ver_todas_solicitudes o solicitudes.gestionar_solicitudes
    Filtros: Estado, tipo, fechas, búsqueda
    """
    model = Solicitud
    template_name = 'solicitudes/lista_solicitudes.html'
    context_object_name = 'solicitudes'
    paginate_by = 25
    filter_form_class = FiltroSolicitudesForm

    def get_queryset(self) -> QuerySet:
        """Retorna solicitudes con relaciones optimizadas y filtros."""
        queryset = super().get_queryset().select_related(
            'tipo_solicitud', 'estado', 'solicitante', 'bodega_origen'
        )
        queryset = scope_solicitudes_for_user(queryset, self.request.user)

        # Aplicar filtros del formulario
        form = self.filter_form_class(self.request.GET)
        if form.is_valid():
            data = form.cleaned_data

            if data.get('estado'):
                queryset = queryset.filter(estado=data['estado'])

            if data.get('tipo'):
                queryset = queryset.filter(tipo_solicitud=data['tipo'])

            if data.get('fecha_desde'):
                queryset = queryset.filter(fecha_solicitud__gte=data['fecha_desde'])

            if data.get('fecha_hasta'):
                queryset = queryset.filter(fecha_solicitud__lte=data['fecha_hasta'])

            if data.get('buscar'):
                q = data['buscar']
                queryset = queryset.filter(
                    Q(numero__icontains=q) |
                    Q(solicitante__username__icontains=q) |
                    Q(area__nombre__icontains=q) |
                    Q(departamento__nombre__icontains=q)
                )

        return queryset.order_by('-fecha_solicitud')

    def get_context_data(self, **kwargs) -> dict:
        """Agrega datos adicionales al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Solicitudes'
        context['form'] = self.filter_form_class(self.request.GET)
        return context


class SolicitudDetailView(ScopedObjectPermissionMixin, BaseAuditedViewMixin, DetailView):
    """
    Vista para ver el detalle de una solicitud.

    Permisos: solicitudes.view_solicitud
    """
    model = Solicitud
    template_name = 'solicitudes/detalle_solicitud.html'
    context_object_name = 'solicitud'
    permission_required = 'solicitudes.view_solicitud'

    def get_queryset(self) -> QuerySet:
        """Optimiza consultas con select_related."""
        return scope_solicitudes_for_user(
            super().get_queryset().select_related(
            'tipo_solicitud', 'estado', 'solicitante', 'aprobador',
            'despachador', 'bodega_origen', 'departamento', 'area'
            ),
            self.request.user
        )

    def has_object_permission(self, obj) -> bool:
        return can_view_solicitud(self.request.user, obj)

    def get_context_data(self, **kwargs) -> dict:
        """Agrega detalles y historial al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Solicitud {self.object.numero}'

        context['detalles'] = self.object.detalles.filter(eliminado=False).select_related(
            'articulo',
            'articulo__categoria',
            'activo',
            'activo__categoria'
        ).order_by('id')

        context['historial'] = self.object.historial.select_related(
            'estado_anterior', 'estado_nuevo', 'usuario'
        )

        # Pasar el origen al contexto para que el template de página completa sepa cuál tabla usar
        context['origen'] = self.request.GET.get('origen', 'mis')

        return context

    def get_template_names(self):
        # Si la petición es AJAX o se solicita modal, devolver el modal según el origen
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest' or self.request.GET.get('modal') == '1':
            if self.request.GET.get('origen') == 'admin':
                return ['solicitudes/partials/modal_detalle_admin.html']
            else:
                return ['solicitudes/partials/modal_detalle_mis_solicitudes.html']
        return [self.template_name]


class MisSolicitudesListView(MisSolicitudesPermissionMixin, BaseAuditedViewMixin, PaginatedListMixin, ListView):
    """
    Vista para ver las solicitudes del usuario actual (MIS SOLICITUDES).

    Permisos: solicitudes.ver_mis_solicitudes
    """
    model = Solicitud
    template_name = 'solicitudes/mis_solicitudes.html'
    context_object_name = 'solicitudes'
    paginate_by = 25

    def get_queryset(self) -> QuerySet:
        """Retorna solo las solicitudes del usuario actual."""
        return super().get_queryset().filter(
            solicitante=self.request.user
        ).select_related(
            'tipo_solicitud', 'estado', 'bodega_origen'
        ).order_by('-fecha_solicitud')

    def get_context_data(self, **kwargs) -> dict:
        """Agrega datos adicionales al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Mis Solicitudes'
        return context


class SolicitudCreateView(BaseAuditedViewMixin, AtomicTransactionMixin, CreateView):
    """
    Vista para crear una nueva solicitud.

    Permisos: solicitudes.add_solicitud
    Auditoría: Registra acción CREAR automáticamente
    Transacción atómica: Garantiza que solicitud y detalles se creen correctamente
    """
    model = Solicitud
    form_class = SolicitudForm
    template_name = 'solicitudes/form_solicitud.html'
    permission_required = 'solicitudes.add_solicitud'

    # Configuración de auditoría
    audit_action = 'CREAR'
    audit_description_template = 'Creó solicitud {obj.numero}'

    # Mensaje de éxito
    success_message = 'Solicitud {obj.numero} creada exitosamente.'

    def get_success_url(self) -> str:
        """Redirige al detalle de la solicitud creada."""
        return reverse_lazy('solicitudes:detalle_solicitud', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs) -> dict:
        """Agrega formset y datos al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Crear Solicitud'
        context['action'] = 'Crear'

        # Las vistas hijas deben sobrescribir esto con el formset correcto
        if self.request.POST:
            context['formset'] = DetalleSolicitudActivoFormSet(self.request.POST)
        else:
            context['formset'] = DetalleSolicitudActivoFormSet()

        return context

    def form_valid(self, form):
        """Procesa el formulario válido usando SolicitudService."""
        context = self.get_context_data()
        formset = context['formset']

        if formset.is_valid():
            solicitud_service = SolicitudService()

            try:
                # Crear solicitud usando service
                self.object = solicitud_service.crear_solicitud(
                    tipo_solicitud=form.cleaned_data['tipo_solicitud'],
                    solicitante=self.request.user,
                    fecha_requerida=form.cleaned_data['fecha_requerida'],
                    motivo=form.cleaned_data['motivo'],
                    titulo_actividad=form.cleaned_data.get('titulo_actividad', ''),
                    objetivo_actividad=form.cleaned_data.get('objetivo_actividad', ''),
                    tipo_choice=form.cleaned_data.get('tipo', 'ARTICULO'),
                    bodega_origen=form.cleaned_data.get('bodega_origen'),
                    departamento=form.cleaned_data.get('departamento'),
                    area=form.cleaned_data.get('area'),
                    equipo=form.cleaned_data.get('equipo'),
                    observaciones=form.cleaned_data.get('observaciones', '')
                )

                # Guardar los detalles
                formset.instance = self.object
                formset.save()

                # Mensaje de éxito y log de auditoría
                messages.success(self.request, self.get_success_message(self.object))
                self.log_action(self.object, self.request)

                return redirect(self.get_success_url())

            except ValidationError as e:
                for field, errors in e.message_dict.items():
                    for error in errors:
                        form.add_error(field if field != '__all__' else None, error)
                return self.form_invalid(form)
        else:
            return self.form_invalid(form)


class SolicitudUpdateView(EditarMisSolicitudesPermissionMixin, BaseAuditedViewMixin, AtomicTransactionMixin, UpdateView):
    """
    Vista para editar una solicitud existente (MIS SOLICITUDES).

    Permisos: solicitudes.editar_mis_solicitudes (si es el solicitante)
             o solicitudes.editar_cualquier_solicitud (si es gestor)
    Auditoría: Registra acción EDITAR automáticamente
    Transacción atómica: Garantiza consistencia de datos
    """
    model = Solicitud
    form_class = SolicitudForm
    template_name = 'solicitudes/form_solicitud.html'

    # Configuración de auditoría
    audit_action = 'EDITAR'
    audit_description_template = 'Editó solicitud {obj.numero}'

    # Mensaje de éxito
    success_message = 'Solicitud {obj.numero} actualizada exitosamente.'

    def get_success_url(self) -> str:
        """Redirige al detalle de la solicitud editada."""
        return reverse_lazy('solicitudes:detalle_solicitud', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs) -> dict:
        """Agrega formset y datos al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Editar Solicitud {self.object.numero}'
        context['action'] = 'Editar'
        context['solicitud'] = self.object

        # Seleccionar el formset correcto según el tipo de solicitud
        if self.object.tipo == 'ARTICULO':
            context['tipo'] = 'ARTICULO'
            if self.request.POST:
                context['formset'] = DetalleSolicitudArticuloFormSet(self.request.POST, instance=self.object)
            else:
                context['formset'] = DetalleSolicitudArticuloFormSet(instance=self.object)
        else:  # ACTIVO
            context['tipo'] = 'ACTIVO'
            if self.request.POST:
                context['formset'] = DetalleSolicitudActivoFormSet(self.request.POST, instance=self.object)
            else:
                context['formset'] = DetalleSolicitudActivoFormSet(instance=self.object)

        return context

    def get_template_names(self):
        # Si la petición es AJAX o se solicita modal, devolver plantilla parcial del formulario
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest' or self.request.GET.get('modal') == '1':
            return ['solicitudes/partials/modal_form.html']
        return [self.template_name]

    def form_valid(self, form):
        """Procesa el formulario válido con formset.

        Si la petición es AJAX, devuelve el detalle actualizado renderizado en parcial
        para que el frontend lo reemplace dentro del modal o cierre el modal.
        """
        context = self.get_context_data()
        formset = context['formset']

        if formset.is_valid():
            # Guardar manualmente para controlar la respuesta en AJAX
            form.save()
            formset.save()

            # Log de auditoría
            self.log_action(self.object, self.request)

            # Si es petición AJAX, devolver el detalle parcial para mostrar en modal
            if self.request.headers.get('x-requested-with') == 'XMLHttpRequest' or self.request.GET.get('modal') == '1':
                # Construir contexto usando la DetailView para mantener consistencia
                detalle_view = SolicitudDetailView()
                detalle_view.request = self.request
                detalle_view.object = self.object
                detalle_context = detalle_view.get_context_data()
                # Usar modal diferente según si es mis solicitudes o admin
                es_mis_solicitudes = self.object.solicitante == self.request.user
                modal_template = 'solicitudes/partials/modal_detalle_mis_solicitudes.html' if es_mis_solicitudes else 'solicitudes/partials/modal_detalle_admin.html'
                return render(self.request, modal_template, detalle_context)

            return super().form_valid(form)
        else:
            # Si es AJAX, re-renderizar el formulario con errores para inyectarlo en el modal
            if self.request.headers.get('x-requested-with') == 'XMLHttpRequest' or self.request.GET.get('modal') == '1':
                return render(self.request, 'solicitudes/partials/modal_form.html', context)
            return self.form_invalid(form)


class SolicitudDeleteView(EliminarMisSolicitudesPermissionMixin, BaseAuditedViewMixin, AtomicTransactionMixin, DeleteView):
    """
    Vista para eliminar una solicitud (MIS SOLICITUDES).

    Permisos: solicitudes.eliminar_mis_solicitudes (si es el solicitante)
             o solicitudes.eliminar_cualquier_solicitud (si es gestor)
    Auditoría: Registra acción ELIMINAR automáticamente
    Transacción atómica: Garantiza consistencia
    """
    model = Solicitud
    template_name = 'solicitudes/eliminar_solicitud.html'
    success_url = reverse_lazy('solicitudes:lista_solicitudes')

    # Configuración de auditoría
    audit_action = 'ELIMINAR'
    audit_description_template = 'Eliminó solicitud {obj.numero}'

    # Mensaje de éxito
    success_message = 'Solicitud {obj.numero} eliminada exitosamente.'

    def get_context_data(self, **kwargs) -> dict:
        """Agrega datos al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Eliminar Solicitud {self.object.numero}'
        context['solicitud'] = self.object
        return context

    def delete(self, request, *args, **kwargs):
        """Procesa la eliminación con log de auditoría."""
        self.object = self.get_object()
        numero = self.object.numero

        # Eliminar
        response = super().delete(request, *args, **kwargs)

        # Log de auditoría (usar número guardado)
        registrar_log_auditoria(
            usuario=request.user,
            accion_glosa='ELIMINAR',
            descripcion=f'Eliminó solicitud {numero}',
            request=request
        )

        messages.success(request, f'Solicitud {numero} eliminada exitosamente.')
        return response


# ==================== VISTAS DE WORKFLOW (APROBAR, RECHAZAR, DESPACHAR) ====================

class SolicitudAprobarView(AprobarSolicitudesPermissionMixin, BaseAuditedViewMixin, AtomicTransactionMixin, DetailView):
    """
    Vista para aprobar una solicitud (GESTIÓN).

    Permisos: solicitudes.aprobar_solicitudes
    Auditoría: Registra acción APROBAR automáticamente
    Transacción atómica: Garantiza consistencia del workflow
    """
    model = Solicitud
    template_name = 'solicitudes/aprobar_solicitud.html'
    context_object_name = 'solicitud'

    # Configuración de auditoría
    audit_action = 'APROBAR'
    audit_description_template = 'Aprobó solicitud {obj.numero}'

    def get_context_data(self, **kwargs) -> dict:
        """Agrega formulario y datos al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Aprobar Solicitud {self.object.numero}'

        if self.request.POST:
            context['form'] = AprobarSolicitudForm(self.request.POST, solicitud=self.object)
        else:
            context['form'] = AprobarSolicitudForm(solicitud=self.object)

        return context

    def get_template_names(self):
        """Devuelve el partial si es modal."""
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest' or self.request.GET.get('modal') == '1':
            return ['solicitudes/aprobar_solicitud.html']
        return [self.template_name]

    def get(self, request, *args, **kwargs):
        """Verifica que la solicitud pueda ser aprobada."""
        self.object = self.get_object()

        # Verificar que no esté ya aprobada
        if self.object.aprobador:
            messages.warning(request, 'Esta solicitud ya fue aprobada.')
            return redirect('solicitudes:detalle_solicitud', pk=self.object.pk)

        # Verificar que esté en estado PENDIENTE
        if self.object.estado.codigo != 'PENDIENTE':
            messages.warning(request, f'Esta solicitud está en estado {self.object.estado.nombre} y no puede ser aprobada.')
            return redirect('solicitudes:detalle_solicitud', pk=self.object.pk)

        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        """Procesa la aprobación de la solicitud usando SolicitudService."""
        self.object = self.get_object()
        form = AprobarSolicitudForm(request.POST, solicitud=self.object)

        if form.is_valid():
            solicitud_service = SolicitudService()

            # Preparar detalles aprobados
            detalles_aprobados = []
            for detalle in self.object.detalles.all():
                field_name = f'cantidad_aprobada_{detalle.id}'
                if field_name in form.cleaned_data:
                    detalles_aprobados.append({
                        'detalle_id': detalle.id,
                        'cantidad_aprobada': form.cleaned_data[field_name]
                    })

            try:
                # Aprobar usando service
                self.object = solicitud_service.aprobar_solicitud(
                    solicitud=self.object,
                    aprobador=request.user,
                    detalles_aprobados=detalles_aprobados,
                    notas_aprobacion=form.cleaned_data.get('notas_aprobacion', '')
                )

                # Log de auditoría
                self.log_action(self.object, request)

                messages.success(request, f'Solicitud {self.object.numero} aprobada exitosamente.')
                return redirect('solicitudes:detalle_solicitud', pk=self.object.pk)

            except ValidationError as e:
                error_msg = str(e.message_dict.get('__all__', [e])[0]) if hasattr(e, 'message_dict') else str(e)
                messages.error(request, error_msg)
                return self.render_to_response(self.get_context_data(form=form))

        # Si el formulario no es válido, mostrar errores
        return self.render_to_response(self.get_context_data(form=form))


class SolicitudRechazarView(RechazarSolicitudesPermissionMixin, BaseAuditedViewMixin, AtomicTransactionMixin, DetailView):
    """
    Vista para rechazar una solicitud (GESTIÓN).

    Permisos: solicitudes.rechazar_solicitudes
    Auditoría: Registra acción RECHAZAR automáticamente
    Transacción atómica: Garantiza consistencia del workflow
    """
    model = Solicitud
    template_name = 'solicitudes/rechazar_solicitud.html'
    context_object_name = 'solicitud'

    # Configuración de auditoría
    audit_action = 'RECHAZAR'
    audit_description_template = 'Rechazó solicitud {obj.numero}'

    def get_template_names(self):
        """Devuelve el partial si es modal."""
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest' or self.request.GET.get('modal') == '1':
            return ['solicitudes/rechazar_solicitud.html']
        return [self.template_name]

    def get_context_data(self, **kwargs) -> dict:
        """Agrega formulario al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Rechazar Solicitud {self.object.numero}'
        context['form'] = RechazarSolicitudForm()
        return context

    def get(self, request, *args, **kwargs):
        """Verifica que la solicitud pueda ser rechazada."""
        self.object = self.get_object()

        # Verificar que no esté finalizada
        if self.object.estado.es_final:
            messages.warning(request, 'No se puede rechazar una solicitud finalizada.')
            return redirect('solicitudes:detalle_solicitud', pk=self.object.pk)

        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        """Procesa el rechazo de la solicitud."""
        self.object = self.get_object()
        form = RechazarSolicitudForm(request.POST)

        if form.is_valid():
            try:
                solicitud_service = SolicitudService()
                
                # Rechazar usando service
                self.object = solicitud_service.rechazar_solicitud(
                    solicitud=self.object,
                    rechazador=request.user,
                    motivo_rechazo=form.cleaned_data['motivo_rechazo']
                )

                # Log de auditoría
                self.log_action(self.object, request)

                messages.success(request, f'Solicitud {self.object.numero} rechazada exitosamente.')
                
                # Si es petición AJAX/modal, devolver el partial del detalle
                if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.GET.get('modal') == '1':
                    return self._render_modal_detail_response(self.object)

                return redirect('solicitudes:detalle_solicitud', pk=self.object.pk)

            except ValidationError as e:
                error_msg = str(e.message_dict.get('__all__', [e])[0]) if hasattr(e, 'message_dict') else str(e)
                messages.error(request, error_msg)
                return self.render_to_response(self.get_context_data(form=form))

        # Si el formulario no es válido, mostrar errores
        return self.render_to_response(self.get_context_data(form=form))

    def _render_modal_detail_response(self, solicitud):
        """Helper para renderizar la respuesta del modal tras acción exitosa."""
        from .repositories import HistorialSolicitudRepository
        historial_repo = HistorialSolicitudRepository()
        context = {
            'solicitud': solicitud,
            'detalles': solicitud.detalles.filter(eliminado=False).select_related(
                'articulo', 'articulo__categoria', 'activo', 'activo__categoria'
            ).order_by('id'),
            'historial': historial_repo.filter_by_solicitud(solicitud)
        }
        es_mis_solicitudes = solicitud.solicitante == self.request.user
        modal_template = 'solicitudes/partials/modal_detalle_mis_solicitudes.html' if es_mis_solicitudes else 'solicitudes/partials/modal_detalle_admin.html'
        return render(self.request, modal_template, context)


class SolicitudDespacharView(DespacharSolicitudesPermissionMixin, BaseAuditedViewMixin, AtomicTransactionMixin, DetailView):
    """
    Vista para despachar una solicitud (GESTIÓN).

    Permisos: solicitudes.despachar_solicitudes
    Auditoría: Registra acción DESPACHAR automáticamente
    Transacción atómica: Garantiza consistencia del workflow
    """
    model = Solicitud
    template_name = 'solicitudes/despachar_solicitud.html'
    context_object_name = 'solicitud'

    def get_queryset(self):
        """Optimiza la consulta para incluir detalles y stock."""
        from django.db.models import Prefetch
        return super().get_queryset().prefetch_related(
            Prefetch(
                'detalles',
                queryset=DetalleSolicitud.objects.select_related(
                    'articulo', 
                    'articulo__unidad_medida', 
                    'activo',
                    'activo__estado'
                )
            )
        )

    # Configuración de auditoría
    audit_action = 'DESPACHAR'
    audit_description_template = 'Despachó solicitud {obj.numero}'

    def get_template_names(self):
        """Devuelve el partial si es modal."""
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest' or self.request.GET.get('modal') == '1':
            return ['solicitudes/despachar_solicitud.html']
        return [self.template_name]

    def get_context_data(self, **kwargs) -> dict:
        """Agrega formulario al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Despachar Solicitud {self.object.numero}'
        
        if self.request.POST:
            context['form'] = DespacharSolicitudForm(self.request.POST, solicitud=self.object)
        else:
            context['form'] = DespacharSolicitudForm(solicitud=self.object)
            
        return context

    def get(self, request, *args, **kwargs):
        """Verifica que la solicitud pueda ser despachada."""
        self.object = self.get_object()

        # Verificar que no esté finalizada
        if self.object.estado.es_final:
            messages.warning(request, 'No se puede despachar una solicitud finalizada.')
            return redirect('solicitudes:detalle_solicitud', pk=self.object.pk)

        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        """Procesa el despacho de la solicitud."""
        self.object = self.get_object()
        form = DespacharSolicitudForm(request.POST, solicitud=self.object)

        if form.is_valid():
            try:
                solicitud_service = SolicitudService()

                # Preparar detalles despachados
                detalles_despachados = []
                for detalle in self.object.detalles.all():
                    field_name = f'cantidad_despachada_{detalle.id}'
                    if field_name in form.cleaned_data:
                        detalles_despachados.append({
                            'detalle_id': detalle.id,
                            'cantidad_despachada': form.cleaned_data[field_name]
                        })

                # Despachar usando service
                self.object = solicitud_service.despachar_solicitud(
                    solicitud=self.object,
                    despachador=request.user,
                    detalles_despachados=detalles_despachados,
                    notas_despacho=form.cleaned_data.get('notas_despacho', '')
                )

                # Log de auditoría
                self.log_action(self.object, request)

                messages.success(request, f'Solicitud {self.object.numero} despachada exitosamente.')
                
                # Si es petición AJAX/modal, devolver parcial
                if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.GET.get('modal') == '1':
                    # Reutilizar el helper de RechazarView
                    rechazar_view = SolicitudRechazarView()
                    rechazar_view.request = request
                    return rechazar_view._render_modal_detail_response(self.object)

                return redirect('solicitudes:detalle_solicitud', pk=self.object.pk)

            except ValidationError as e:
                error_msg = str(e.message_dict.get('__all__', [e])[0]) if hasattr(e, 'message_dict') else str(e)
                messages.error(request, error_msg)
                return self.render_to_response(self.get_context_data(form=form))

        # Si el formulario no es válido
        return self.render_to_response(self.get_context_data(form=form))


class SolicitudComprarView(DespacharSolicitudesPermissionMixin, BaseAuditedViewMixin, AtomicTransactionMixin, View):
    """
    Vista para marcar una solicitud como comprada (en proceso de compra).

    Permisos: solicitudes.despachar_solicitudes (mismo permiso que despachar)
    Auditoría: Registra acción COMPRAR automáticamente
    Transacción atómica: Garantiza consistencia del workflow
    """

    # Configuración de auditoría
    audit_action = 'COMPRAR'
    audit_description_template = 'Envió a compras la solicitud {obj.numero}'

    def post(self, request, pk):
        """Procesa el envío de la solicitud a compras usando SolicitudService."""
        try:
            solicitud = Solicitud.objects.get(pk=pk)
        except Solicitud.DoesNotExist:
            messages.error(request, 'Solicitud no encontrada.')
            return redirect('solicitudes:lista_solicitudes')

        # Verificar que no esté finalizada
        if solicitud.estado.es_final:
            messages.warning(request, 'No se puede enviar a compras una solicitud finalizada.')
            return redirect('solicitudes:detalle_solicitud', pk=solicitud.pk)

        try:
            solicitud_service = SolicitudService()

            # Marcar como comprada usando service
            solicitud = solicitud_service.comprar_solicitud(
                solicitud=solicitud,
                comprador=request.user,
                notas_compra=request.POST.get('notas_compra', '')
            )

            # Log de auditoría
            self.log_action(solicitud, request)

            messages.success(request, f'Solicitud {solicitud.numero} enviada a compras exitosamente.')

        except ValidationError as e:
            error_msg = str(e.message_dict.get('__all__', [e])[0]) if hasattr(e, 'message_dict') else str(e)
            messages.error(request, error_msg)

        # Si es petición AJAX/modal, devolver el partial actualizado
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.GET.get('modal') == '1':
            from .repositories import HistorialSolicitudRepository
            historial_repo = HistorialSolicitudRepository()
            context = {
                'solicitud': solicitud,
                'detalles': solicitud.detalles.filter(eliminado=False).select_related(
                    'articulo', 'articulo__categoria', 'activo', 'activo__categoria'
                ).order_by('id'),
                'historial': historial_repo.filter_by_solicitud(solicitud)
            }
            # Usar modal diferente según si es mis solicitudes o admin
            es_mis_solicitudes = solicitud.solicitante == request.user
            modal_template = 'solicitudes/partials/modal_detalle_mis_solicitudes.html' if es_mis_solicitudes else 'solicitudes/partials/modal_detalle_admin.html'
            return render(request, modal_template, context)

        return redirect('solicitudes:detalle_solicitud', pk=solicitud.pk)


# ==================== VISTAS ESPECÍFICAS PARA ACTIVOS ====================

class SolicitudActivoListView(VerSolicitudesBienesPermissionMixin, SolicitudListView):
    """
    Vista para listar solicitudes de activos/bienes (SOLICITUD BIENES).

    Permisos: solicitudes.ver_solicitudes_bienes
    """
    template_name = 'solicitudes/lista_solicitudes_activos.html'

    def get_queryset(self) -> QuerySet:
        """Filtra solo solicitudes de activos."""
        return super().get_queryset().filter(tipo='ACTIVO')

    def get_context_data(self, **kwargs) -> dict:
        """Agrega datos adicionales al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Solicitudes de Activos/Bienes'
        context['tipo'] = 'ACTIVO'
        return context


class SolicitudActivoCreateView(CrearSolicitudBienesPermissionMixin, SolicitudCreateView):
    """
    Vista para crear una nueva solicitud de bienes (SOLICITUD BIENES).

    Permisos: solicitudes.crear_solicitud_bienes
    """
    template_name = 'solicitudes/form_solicitud_bienes.html'

    def get_template_names(self):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest' or self.request.GET.get('modal') == '1':
            return ['solicitudes/partials/modal_form_crear.html']
        return [self.template_name]

    def get_context_data(self, **kwargs) -> dict:
        """Agrega datos adicionales al contexto y lista de activos disponibles."""
        from apps.activos.models import Activo

        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Crear Solicitud de Bienes'
        context['action'] = 'Crear'
        context['tipo'] = 'ACTIVO'

        activos = Activo.objects.filter(
            activo=True, eliminado=False
        ).select_related('categoria').order_by('codigo')

        context['activos'] = activos
        context['items'] = activos
        context['form_action'] = self.request.get_full_path()
        context['show_bodega'] = False
        context['item_type'] = 'bien'
        context['item_type_label'] = 'Bien/Activo'
        context['item_id_field'] = 'activo_id'
        context['table_id'] = 'tabla-bienes'
        context['tbody_id'] = 'tbody-bienes'
        context['empty_state_id'] = 'sin-bienes'
        context['btn_agregar_id'] = 'btn-agregar-bien'
        context['items_title'] = 'Bienes/Activos Solicitados'
        context['has_unidad'] = False
        context['show_stock'] = False
        context['min_step'] = '1'
        context['min_value'] = '1'
        context['empty_message'] = 'No hay bienes/activos agregados. Haga clic en "Agregar Bien/Activo" para comenzar.'
        context['duplicate_message'] = 'Este bien/activo ya ha sido agregado'
        context['validation_message'] = 'Debe agregar al menos un bien/activo a la solicitud'
        context['qty_validation_message'] = 'Todas las cantidades deben ser mayores a cero'

        return context

    def form_valid(self, form):
        """Procesa el formulario válido usando SolicitudService con tipo ACTIVO."""
        from django.db import transaction

        try:
            with transaction.atomic():
                # Crear solicitud usando service con tipo ACTIVO
                solicitud_service = SolicitudService()
                self.object = solicitud_service.crear_solicitud(
                    tipo_solicitud=form.cleaned_data['tipo_solicitud'],
                    solicitante=self.request.user,
                    fecha_requerida=form.cleaned_data['fecha_requerida'],
                    motivo=form.cleaned_data['motivo'],
                    titulo_actividad=form.cleaned_data.get('titulo_actividad', ''),
                    objetivo_actividad=form.cleaned_data.get('objetivo_actividad', ''),
                    tipo_choice='ACTIVO',  # FORZAR TIPO ACTIVO
                    bodega_origen=None,  # Los bienes no tienen bodega
                    departamento=form.cleaned_data.get('departamento'),
                    area=form.cleaned_data.get('area'),
                    equipo=form.cleaned_data.get('equipo'),
                    observaciones=form.cleaned_data.get('observaciones', '')
                )

                # Procesar detalles de bienes desde el POST
                detalles = self._extraer_detalles_post(self.request.POST)

                if not detalles:
                    form.add_error(None, 'Debe agregar al menos un bien/activo a la solicitud')
                    return self.form_invalid(form)

                # Crear detalles de bienes
                for detalle_data in detalles:
                    DetalleSolicitud.objects.create(
                        solicitud=self.object,
                        activo_id=detalle_data['activo_id'],
                        cantidad_solicitada=Decimal(str(detalle_data['cantidad_solicitada'])),
                        observaciones=detalle_data.get('observaciones', '')
                    )

                # Mensaje de éxito y log de auditoría
                messages.success(self.request, self.get_success_message(self.object))
                self.log_action(self.object, self.request)

                # Si es AJAX/modal, devolver detalle parcial
                if self.request.headers.get('x-requested-with') == 'XMLHttpRequest' or self.request.GET.get('modal') == '1':
                    return self._render_modal_detalle()

                return redirect(self.get_success_url())

        except ValidationError as e:
            for field, errors in e.message_dict.items():
                for error in errors:
                    form.add_error(field if field != '__all__' else None, error)
            return self.form_invalid(form)
        except Exception as e:
            form.add_error(None, f'Error al crear la solicitud: {str(e)}')
            return self.form_invalid(form)

    def _render_modal_detalle(self):
        """Renderiza el detalle de la solicitud recién creada para mostrar en modal."""
        context = {
            'solicitud': self.object,
            'detalles': self.object.detalles.filter(eliminado=False).select_related(
                'articulo', 'articulo__categoria', 'activo', 'activo__categoria'
            ).order_by('id'),
            'historial': self.object.historial.select_related(
                'estado_anterior', 'estado_nuevo', 'usuario'
            )
        }
        # Usar modal diferente según si es mis solicitudes o admin
        # Las solicitudes creadas siempre son "mis solicitudes" para el usuario actual
        modal_template = 'solicitudes/partials/modal_detalle_mis_solicitudes.html'
        return render(self.request, modal_template, context)

    def _extraer_detalles_post(self, post_data):
        """
        Extrae los detalles de bienes del POST.
        Formato esperado: detalles[0][activo_id], detalles[0][cantidad_solicitada], etc.
        """
        detalles = []
        indices = set()

        # Identificar todos los índices presentes
        for key in post_data.keys():
            if key.startswith('detalles['):
                # Extraer índice: detalles[0][campo] -> 0
                indice = key.split('[')[1].split(']')[0]
                indices.add(indice)

        # Extraer datos para cada índice
        for indice in indices:
            activo_id = post_data.get(f'detalles[{indice}][activo_id]')
            cantidad_solicitada = post_data.get(f'detalles[{indice}][cantidad_solicitada]')

            if activo_id and cantidad_solicitada:
                detalle = {
                    'activo_id': int(activo_id),
                    'cantidad_solicitada': float(cantidad_solicitada),
                    'observaciones': post_data.get(f'detalles[{indice}][observaciones]', '')
                }

                detalles.append(detalle)

        return detalles


class SolicitudActivoUpdateView(SolicitudUpdateView):
    """Vista para editar una solicitud de bienes."""
    template_name = 'solicitudes/form_solicitud.html'

    def get_queryset(self) -> QuerySet:
        """Solo solicitudes de bienes."""
        return super().get_queryset().filter(tipo='ACTIVO')

    def get(self, request, *args, **kwargs):
        """Verifica que la solicitud pueda ser editada."""
        self.object = self.get_object()

        # Solo se pueden editar si están en estado pendiente o rechazada
        if self.object.estado.codigo not in ['PENDIENTE', 'RECHAZADA']:
            messages.error(request, 'Solo se pueden editar solicitudes en estado Pendiente o Rechazada.')
            return redirect('solicitudes:detalle_solicitud', pk=self.object.pk)

        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs) -> dict:
        """Agrega datos adicionales al contexto y formset de BIENES."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Editar Solicitud de Bienes {self.object.numero}'
        context['action'] = 'Editar'
        context['tipo'] = 'ACTIVO'
        context['solicitud'] = self.object

        # Usar formset de BIENES/ACTIVOS
        if self.request.POST:
            context['formset'] = DetalleSolicitudActivoFormSet(self.request.POST, instance=self.object)
        else:
            context['formset'] = DetalleSolicitudActivoFormSet(instance=self.object)

        return context

    def form_valid(self, form):
        """Asegura que no tenga bodega."""
        form.instance.bodega_origen = None
        return super().form_valid(form)


# ==================== VISTAS ESPECÍFICAS PARA ARTÍCULOS ====================

class SolicitudArticuloListView(VerSolicitudesArticulosPermissionMixin, SolicitudListView):
    """
    Vista para listar solicitudes de artículos (SOLICITUD ARTÍCULOS).

    Permisos: solicitudes.ver_solicitudes_articulos
    """
    template_name = 'solicitudes/lista_solicitudes_articulos.html'

    def get_queryset(self) -> QuerySet:
        """Filtra solo solicitudes de artículos."""
        return super().get_queryset().filter(tipo='ARTICULO')

    def get_context_data(self, **kwargs) -> dict:
        """Agrega datos adicionales al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Solicitudes de Artículos'
        context['tipo'] = 'ARTICULO'
        return context


class SolicitudArticuloCreateView(CrearSolicitudArticulosPermissionMixin, SolicitudCreateView):
    """
    Vista para crear una nueva solicitud de artículos (SOLICITUD ARTÍCULOS).

    Permisos: solicitudes.crear_solicitud_articulos
    """
    template_name = 'solicitudes/form_solicitud_articulos.html'

    def get_template_names(self):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest' or self.request.GET.get('modal') == '1':
            return ['solicitudes/partials/modal_form_crear.html']
        return [self.template_name]

    def get_context_data(self, **kwargs) -> dict:
        """Agrega datos adicionales al contexto y lista de artículos disponibles."""
        from apps.bodega.models import Articulo

        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Crear Solicitud de Artículos'
        context['action'] = 'Crear'
        context['tipo'] = 'ARTICULO'

        articulos = Articulo.objects.filter(
            activo=True, eliminado=False
        ).select_related('categoria', 'unidad_medida').order_by('codigo')

        context['articulos'] = articulos
        context['items'] = articulos
        context['form_action'] = self.request.get_full_path()
        context['show_bodega'] = True
        context['item_type'] = 'articulo'
        context['item_type_label'] = 'Artículo'
        context['item_id_field'] = 'articulo_id'
        context['table_id'] = 'tabla-articulos'
        context['tbody_id'] = 'tbody-articulos'
        context['empty_state_id'] = 'sin-articulos'
        context['btn_agregar_id'] = 'btn-agregar-articulo'
        context['items_title'] = 'Artículos Solicitados'
        context['has_unidad'] = True
        context['show_stock'] = True
        context['min_step'] = '0.01'
        context['min_value'] = '0.01'
        context['empty_message'] = 'No hay artículos agregados. Haga clic en "Agregar Artículo" para comenzar.'
        context['duplicate_message'] = 'Este artículo ya ha sido agregado'
        context['validation_message'] = 'Debe agregar al menos un artículo a la solicitud'
        context['qty_validation_message'] = 'Todas las cantidades deben ser mayores a cero'

        return context

    def form_valid(self, form):
        """Procesa el formulario válido usando SolicitudService con tipo ARTICULO."""
        from django.db import transaction

        try:
            with transaction.atomic():
                # Crear solicitud usando service con tipo ARTICULO
                solicitud_service = SolicitudService()
                self.object = solicitud_service.crear_solicitud(
                    tipo_solicitud=form.cleaned_data['tipo_solicitud'],
                    solicitante=self.request.user,
                    fecha_requerida=form.cleaned_data['fecha_requerida'],
                    motivo=form.cleaned_data['motivo'],
                    titulo_actividad=form.cleaned_data.get('titulo_actividad', ''),
                    objetivo_actividad=form.cleaned_data.get('objetivo_actividad', ''),
                    tipo_choice='ARTICULO',  # FORZAR TIPO ARTICULO
                    bodega_origen=form.cleaned_data.get('bodega_origen'),
                    departamento=form.cleaned_data.get('departamento'),
                    area=form.cleaned_data.get('area'),
                    equipo=form.cleaned_data.get('equipo'),
                    observaciones=form.cleaned_data.get('observaciones', '')
                )

                # Procesar detalles de artículos desde el POST
                detalles = self._extraer_detalles_post(self.request.POST)

                if not detalles:
                    form.add_error(None, 'Debe agregar al menos un artículo a la solicitud')
                    return self.form_invalid(form)

                # Crear detalles de artículos
                for detalle_data in detalles:
                    DetalleSolicitud.objects.create(
                        solicitud=self.object,
                        articulo_id=detalle_data['articulo_id'],
                        cantidad_solicitada=Decimal(str(detalle_data['cantidad_solicitada'])),
                        observaciones=detalle_data.get('observaciones', '')
                    )

                # Mensaje de éxito y log de auditoría
                messages.success(self.request, self.get_success_message(self.object))
                self.log_action(self.object, self.request)

                # Si es AJAX/modal, devolver detalle parcial
                if self.request.headers.get('x-requested-with') == 'XMLHttpRequest' or self.request.GET.get('modal') == '1':
                    return self._render_modal_detalle()

                return redirect(self.get_success_url())

        except ValidationError as e:
            for field, errors in e.message_dict.items():
                for error in errors:
                    form.add_error(field if field != '__all__' else None, error)
            return self.form_invalid(form)
        except Exception as e:
            form.add_error(None, f'Error al crear la solicitud: {str(e)}')
            return self.form_invalid(form)

    def form_invalid(self, form):
        print(f"================ FORM ERRORS ================\n{form.errors}\nNON_FIELD_ERRORS: {form.non_field_errors()}\n=============================================")
        return super().form_invalid(form)

    def _render_modal_detalle(self):
        """Renderiza el detalle de la solicitud recién creada para mostrar en modal."""
        context = {
            'solicitud': self.object,
            'detalles': self.object.detalles.filter(eliminado=False).select_related(
                'articulo', 'articulo__categoria', 'activo', 'activo__categoria'
            ).order_by('id'),
            'historial': self.object.historial.select_related(
                'estado_anterior', 'estado_nuevo', 'usuario'
            )
        }
        # Usar modal diferente según si es mis solicitudes o admin
        # Las solicitudes creadas siempre son "mis solicitudes" para el usuario actual
        modal_template = 'solicitudes/partials/modal_detalle_mis_solicitudes.html'
        return render(self.request, modal_template, context)

    def _extraer_detalles_post(self, post_data):
        """
        Extrae los detalles de artículos del POST.
        Formato esperado: detalles[0][articulo_id], detalles[0][cantidad_solicitada], etc.
        """
        detalles = []
        indices = set()

        # Identificar todos los índices presentes
        for key in post_data.keys():
            if key.startswith('detalles['):
                # Extraer índice: detalles[0][campo] -> 0
                indice = key.split('[')[1].split(']')[0]
                indices.add(indice)

        # Extraer datos para cada índice
        for indice in indices:
            articulo_id = post_data.get(f'detalles[{indice}][articulo_id]')
            cantidad_solicitada = post_data.get(f'detalles[{indice}][cantidad_solicitada]')

            if articulo_id and cantidad_solicitada:
                detalle = {
                    'articulo_id': int(articulo_id),
                    'cantidad_solicitada': float(cantidad_solicitada),
                    'observaciones': post_data.get(f'detalles[{indice}][observaciones]', '')
                }

                detalles.append(detalle)

        return detalles


class SolicitudArticuloUpdateView(SolicitudUpdateView):
    """Vista para editar una solicitud de artículos."""
    template_name = 'solicitudes/form_solicitud.html'

    def get_queryset(self) -> QuerySet:
        """Solo solicitudes de artículos."""
        return super().get_queryset().filter(tipo='ARTICULO')

    def get(self, request, *args, **kwargs):
        """Verifica que la solicitud pueda ser editada."""
        self.object = self.get_object()

        # Solo se pueden editar si están en estado pendiente o rechazada
        if self.object.estado.codigo not in ['PENDIENTE', 'RECHAZADA']:
            messages.error(request, 'Solo se pueden editar solicitudes en estado Pendiente o Rechazada.')
            return redirect('solicitudes:detalle_solicitud', pk=self.object.pk)

        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs) -> dict:
        """Agrega datos adicionales al contexto y formset de ARTÍCULOS."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Editar Solicitud de Artículos {self.object.numero}'
        context['action'] = 'Editar'
        context['tipo'] = 'ARTICULO'
        context['solicitud'] = self.object

        # Usar formset de ARTÍCULOS
        if self.request.POST:
            context['formset'] = DetalleSolicitudArticuloFormSet(self.request.POST, instance=self.object)
        else:
            context['formset'] = DetalleSolicitudArticuloFormSet(instance=self.object)

        return context


# ==================== VISTAS MANTENEDORES: TIPOS DE SOLICITUD ====================


class TipoSolicitudListView(BaseAuditedViewMixin, PaginatedListMixin, ListView):
    """
    Vista para listar tipos de solicitud.

    Permisos: solicitudes.view_tiposolicitud
    Utiliza: TipoSolicitudRepository para acceso a datos optimizado
    """
    model = TipoSolicitud
    template_name = 'solicitudes/mantenedores/tipo_solicitud/lista.html'
    context_object_name = 'tipos'
    permission_required = 'solicitudes.view_tiposolicitud'
    paginate_by = 25

    def get_queryset(self) -> QuerySet:
        """Retorna tipos de solicitud usando repository."""
        from .repositories import TipoSolicitudRepository
        tipo_repo = TipoSolicitudRepository()

        # Incluir inactivos y eliminados para administración
        queryset = TipoSolicitud.objects.filter(eliminado=False).order_by('codigo')

        # Búsqueda por query string
        query = self.request.GET.get('q', '').strip()
        if query:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(codigo__icontains=query) |
                Q(nombre__icontains=query)
            )

        return queryset

    def get_context_data(self, **kwargs) -> dict:
        """Agrega datos adicionales al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Tipos de Solicitud'
        context['puede_crear'] = self.request.user.has_perm('solicitudes.add_tiposolicitud')
        return context


class TipoSolicitudCreateView(BaseAuditedViewMixin, CreateView):
    """
    Vista para crear un nuevo tipo de solicitud.

    Permisos: solicitudes.add_tiposolicitud
    Auditoría: Registra acción CREAR automáticamente
    """
    model = TipoSolicitud
    form_class = TipoSolicitudForm
    template_name = 'solicitudes/mantenedores/tipo_solicitud/form.html'
    permission_required = 'solicitudes.add_tiposolicitud'
    success_url = reverse_lazy('solicitudes:tipo_solicitud_lista')

    # Configuración de auditoría
    audit_action = 'CREAR'
    audit_description_template = 'Creó tipo de solicitud {obj.codigo} - {obj.nombre}'

    # Mensaje de éxito
    success_message = 'Tipo de solicitud {obj.nombre} creado exitosamente.'

    def get_context_data(self, **kwargs) -> dict:
        """Agrega datos al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Crear Tipo de Solicitud'
        context['action'] = 'Crear'
        return context

    def form_valid(self, form):
        """Procesa el formulario válido con log de auditoría."""
        response = super().form_valid(form)
        messages.success(self.request, self.get_success_message(self.object))
        self.log_action(self.object, self.request)
        return response


class TipoSolicitudUpdateView(BaseAuditedViewMixin, UpdateView):
    """
    Vista para editar un tipo de solicitud existente. Soporta modo modal (AJAX).

    Permisos: solicitudes.change_tiposolicitud
    Auditoría: Registra acción EDITAR automáticamente
    """
    model = TipoSolicitud
    form_class = TipoSolicitudForm
    template_name = 'solicitudes/mantenedores/tipo_solicitud/form.html'
    permission_required = 'solicitudes.change_tiposolicitud'
    success_url = reverse_lazy('solicitudes:menu_solicitudes')

    # Configuración de auditoría
    audit_action = 'EDITAR'
    audit_description_template = 'Editó tipo de solicitud {obj.codigo} - {obj.nombre}'

    # Mensaje de éxito
    success_message = 'Tipo de solicitud {obj.nombre} actualizado exitosamente.'

    def get_queryset(self) -> QuerySet:
        """Solo permite editar tipos no eliminados."""
        return super().get_queryset().filter(eliminado=False)

    def get_template_names(self):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return ['solicitudes/mantenedores/tipo_solicitud/modal_editar.html']
        return [self.template_name]

    def get_context_data(self, **kwargs) -> dict:
        """Agrega datos al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Editar Tipo de Solicitud: {self.object.nombre}'
        context['action'] = 'Actualizar'
        context['tipo'] = self.object
        return context

    def form_valid(self, form):
        """Procesa el formulario válido con log de auditoría."""
        response = super().form_valid(form)
        messages.success(self.request, self.get_success_message(self.object))
        self.log_action(self.object, self.request)
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': True})
        return response

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.shortcuts import render as django_render
            return django_render(
                self.request,
                'solicitudes/mantenedores/tipo_solicitud/modal_editar.html',
                self.get_context_data(form=form)
            )
        return super().form_invalid(form)


class TipoSolicitudDeleteView(BaseAuditedViewMixin, DeleteView):
    """
    Vista para eliminar (soft delete) un tipo de solicitud.

    Permisos: solicitudes.delete_tiposolicitud
    Auditoría: Registra acción ELIMINAR automáticamente
    """
    model = TipoSolicitud
    template_name = 'solicitudes/mantenedores/tipo_solicitud/eliminar.html'
    permission_required = 'solicitudes.delete_tiposolicitud'
    success_url = reverse_lazy('solicitudes:tipo_solicitud_lista')

    # Configuración de auditoría
    audit_action = 'ELIMINAR'
    audit_description_template = 'Eliminó tipo de solicitud {obj.codigo} - {obj.nombre}'

    # Mensaje de éxito
    success_message = 'Tipo de solicitud {obj.nombre} eliminado exitosamente.'

    def get_queryset(self) -> QuerySet:
        """Solo permite eliminar tipos no eliminados."""
        return super().get_queryset().filter(eliminado=False)

    def get_context_data(self, **kwargs) -> dict:
        """Agrega datos al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Eliminar Tipo de Solicitud: {self.object.nombre}'
        context['tipo'] = self.object

        # Verificar si hay solicitudes asociadas
        context['tiene_solicitudes'] = self.object.solicitudes.filter(eliminado=False).exists()
        context['count_solicitudes'] = self.object.solicitudes.filter(eliminado=False).count()

        return context

    def delete(self, request, *args, **kwargs):
        """Elimina usando soft delete."""
        self.object = self.get_object()

        # Verificar si tiene solicitudes asociadas
        if self.object.solicitudes.filter(eliminado=False).exists():
            messages.error(
                request,
                f'No se puede eliminar el tipo "{self.object.nombre}" porque tiene solicitudes asociadas. '
                'Desactívelo en su lugar.'
            )
            return redirect('solicitudes:tipo_solicitud_lista')

        # Soft delete
        self.object.eliminado = True
        self.object.activo = False
        self.object.save()

        messages.success(request, self.get_success_message(self.object))
        self.log_action(self.object, request)

        return redirect(self.success_url)


# ==================== VISTAS MANTENEDORES: ESTADOS DE SOLICITUD ====================


class EstadoSolicitudListView(BaseAuditedViewMixin, PaginatedListMixin, ListView):
    """
    Vista para listar estados de solicitud.

    Permisos: solicitudes.view_estadosolicitud
    Utiliza: EstadoSolicitudRepository para acceso a datos optimizado
    """
    model = EstadoSolicitud
    template_name = 'solicitudes/mantenedores/estado_solicitud/lista.html'
    context_object_name = 'estados'
    permission_required = 'solicitudes.view_estadosolicitud'
    paginate_by = 25

    def get_queryset(self) -> QuerySet:
        """Retorna estados de solicitud usando repository."""
        from .repositories import EstadoSolicitudRepository
        estado_repo = EstadoSolicitudRepository()

        # Incluir inactivos y eliminados para administración
        queryset = EstadoSolicitud.objects.filter(eliminado=False).order_by('codigo')

        # Búsqueda por query string
        query = self.request.GET.get('q', '').strip()
        if query:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(codigo__icontains=query) |
                Q(nombre__icontains=query)
            )

        return queryset

    def get_context_data(self, **kwargs) -> dict:
        """Agrega datos adicionales al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Estados de Solicitud'
        context['puede_crear'] = self.request.user.has_perm('solicitudes.add_estadosolicitud')
        return context


class EstadoSolicitudCreateView(BaseAuditedViewMixin, CreateView):
    """
    Vista para crear un nuevo estado de solicitud.

    Permisos: solicitudes.add_estadosolicitud
    Auditoría: Registra acción CREAR automáticamente
    """
    model = EstadoSolicitud
    form_class = EstadoSolicitudForm
    template_name = 'solicitudes/mantenedores/estado_solicitud/form.html'
    permission_required = 'solicitudes.add_estadosolicitud'
    success_url = reverse_lazy('solicitudes:estado_solicitud_lista')

    # Configuración de auditoría
    audit_action = 'CREAR'
    audit_description_template = 'Creó estado de solicitud {obj.codigo} - {obj.nombre}'

    # Mensaje de éxito
    success_message = 'Estado de solicitud {obj.nombre} creado exitosamente.'

    def get_context_data(self, **kwargs) -> dict:
        """Agrega datos al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Crear Estado de Solicitud'
        context['action'] = 'Crear'
        return context

    def form_valid(self, form):
        """Procesa el formulario válido con log de auditoría."""
        response = super().form_valid(form)
        messages.success(self.request, self.get_success_message(self.object))
        self.log_action(self.object, self.request)
        return response


class EstadoSolicitudUpdateView(BaseAuditedViewMixin, UpdateView):
    """
    Vista para editar un estado de solicitud existente. Soporta modo modal (AJAX).

    Permisos: solicitudes.change_estadosolicitud
    Auditoría: Registra acción EDITAR automáticamente
    """
    model = EstadoSolicitud
    form_class = EstadoSolicitudForm
    template_name = 'solicitudes/mantenedores/estado_solicitud/form.html'
    permission_required = 'solicitudes.change_estadosolicitud'
    success_url = reverse_lazy('solicitudes:menu_solicitudes')

    # Configuración de auditoría
    audit_action = 'EDITAR'
    audit_description_template = 'Editó estado de solicitud {obj.codigo} - {obj.nombre}'

    # Mensaje de éxito
    success_message = 'Estado de solicitud {obj.nombre} actualizado exitosamente.'

    def get_queryset(self) -> QuerySet:
        """Solo permite editar estados no eliminados."""
        return super().get_queryset().filter(eliminado=False)

    def get_template_names(self):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return ['solicitudes/mantenedores/estado_solicitud/modal_editar.html']
        return [self.template_name]

    def get_context_data(self, **kwargs) -> dict:
        """Agrega datos al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Editar Estado de Solicitud: {self.object.nombre}'
        context['action'] = 'Actualizar'
        context['estado'] = self.object
        return context

    def form_valid(self, form):
        """Procesa el formulario válido con log de auditoría."""
        response = super().form_valid(form)
        messages.success(self.request, self.get_success_message(self.object))
        self.log_action(self.object, self.request)
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': True})
        return response

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.shortcuts import render as django_render
            return django_render(
                self.request,
                'solicitudes/mantenedores/estado_solicitud/modal_editar.html',
                self.get_context_data(form=form)
            )
        return super().form_invalid(form)


class EstadoSolicitudDeleteView(BaseAuditedViewMixin, DeleteView):
    """
    Vista para eliminar (soft delete) un estado de solicitud.

    Permisos: solicitudes.delete_estadosolicitud
    Auditoría: Registra acción ELIMINAR automáticamente
    """
    model = EstadoSolicitud
    template_name = 'solicitudes/mantenedores/estado_solicitud/eliminar.html'
    permission_required = 'solicitudes.delete_estadosolicitud'
    success_url = reverse_lazy('solicitudes:estado_solicitud_lista')

    # Configuración de auditoría
    audit_action = 'ELIMINAR'
    audit_description_template = 'Eliminó estado de solicitud {obj.codigo} - {obj.nombre}'

    # Mensaje de éxito
    success_message = 'Estado de solicitud {obj.nombre} eliminado exitosamente.'

    def get_queryset(self) -> QuerySet:
        """Solo permite eliminar estados no eliminados."""
        return super().get_queryset().filter(eliminado=False)

    def get_context_data(self, **kwargs) -> dict:
        """Agrega datos al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Eliminar Estado de Solicitud: {self.object.nombre}'
        context['estado'] = self.object

        # Verificar si hay solicitudes asociadas
        context['tiene_solicitudes'] = self.object.solicitudes.filter(eliminado=False).exists()
        context['count_solicitudes'] = self.object.solicitudes.filter(eliminado=False).count()

        return context

    def delete(self, request, *args, **kwargs):
        """Elimina usando soft delete."""
        self.object = self.get_object()

        # Verificar si tiene solicitudes asociadas
        if self.object.solicitudes.filter(eliminado=False).exists():
            messages.error(
                request,
                f'No se puede eliminar el estado "{self.object.nombre}" porque tiene solicitudes asociadas. '
                'Desactívelo en su lugar.'
            )
            return redirect('solicitudes:estado_solicitud_lista')

        # Soft delete
        self.object.eliminado = True
        self.object.activo = False
        self.object.save()

        messages.success(request, self.get_success_message(self.object))
        self.log_action(self.object, request)

        return redirect(self.success_url)


# ==================== IMPORTACION EXCEL PARA MANTENEDORES ====================

from apps.bodega.excel_services.importacion_excel import ImportacionExcelService
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required


@login_required
def tipo_solicitud_descargar_plantilla(request):
    """Vista para descargar plantilla Excel de tipos de solicitud."""
    contenido = ImportacionExcelService.generar_plantilla_tipos_solicitud()
    response = HttpResponse(contenido, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="plantilla_tipos_solicitud.xlsx"'
    return response


@login_required
def tipo_solicitud_importar_excel(request):
    """Vista para importar tipos de solicitud desde Excel."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Metodo no permitido'}, status=405)
    if 'archivo' not in request.FILES:
        return JsonResponse({'error': 'No se proporciono archivo'}, status=400)
    archivo = request.FILES['archivo']
    es_valido, mensaje_error = ImportacionExcelService.validar_archivo_excel(archivo)
    if not es_valido:
        return JsonResponse({'error': mensaje_error}, status=400)
    try:
        creadas, actualizadas, errores = ImportacionExcelService.importar_tipos_solicitud(archivo, request.user)
        mensaje = f"Importacion completada: {creadas} tipos creados, {actualizadas} actualizados"
        if errores:
            mensaje += f". Errores: {len(errores)}"
        return JsonResponse({
            'success': True,
            'mensaje': mensaje,
            'creadas': creadas,
            'actualizadas': actualizadas,
            'errores': errores[:10]
        })
    except Exception as e:
        return JsonResponse({'error': f'Error al importar: {str(e)}'}, status=500)


@login_required
def estado_solicitud_descargar_plantilla(request):
    """Vista para descargar plantilla Excel de estados de solicitud."""
    contenido = ImportacionExcelService.generar_plantilla_estados_solicitud()
    response = HttpResponse(contenido, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="plantilla_estados_solicitud.xlsx"'
    return response


@login_required
def estado_solicitud_importar_excel(request):
    """Vista para importar estados de solicitud desde Excel."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Metodo no permitido'}, status=405)
    if 'archivo' not in request.FILES:
        return JsonResponse({'error': 'No se proporciono archivo'}, status=400)
    archivo = request.FILES['archivo']
    es_valido, mensaje_error = ImportacionExcelService.validar_archivo_excel(archivo)
    if not es_valido:
        return JsonResponse({'error': mensaje_error}, status=400)
    try:
        creadas, actualizadas, errores = ImportacionExcelService.importar_estados_solicitud(archivo, request.user)
        mensaje = f"Importacion completada: {creadas} estados creados, {actualizadas} actualizados"
        if errores:
            mensaje += f". Errores: {len(errores)}"
        return JsonResponse({
            'success': True,
            'mensaje': mensaje,
            'creadas': creadas,
            'actualizadas': actualizadas,
            'errores': errores[:10]
        })
    except Exception as e:
        return JsonResponse({'error': f'Error al importar: {str(e)}'}, status=500)


# ==================== EXPORTAR DATOS MAESTROS (GESTORES) ====================

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO


def _excel_response(contenido: bytes, filename: str):
    resp = HttpResponse(contenido, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


@login_required
def tipo_solicitud_exportar_excel(request):
    """Exporta el listado de tipos de solicitud a Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tipos de Solicitud"

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center")

    headers = ["Código", "Nombre", "Descripción", "Requiere Aprobación", "Activo"]
    widths  = [15,        35,       45,             22,                    10]

    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"

    odd_fill = PatternFill("solid", fgColor="EBF0F8")
    for row_idx, obj in enumerate(
        TipoSolicitud.objects.filter(eliminado=False).order_by('codigo'), start=2
    ):
        fill = odd_fill if row_idx % 2 == 0 else None
        fila = [
            obj.codigo,
            obj.nombre,
            obj.descripcion or "",
            "SI" if obj.requiere_aprobacion else "NO",
            "SI" if obj.activo else "NO",
        ]
        for col, valor in enumerate(fila, start=1):
            cell = ws.cell(row=row_idx, column=col, value=valor)
            cell.alignment = center if col in (4, 5) else Alignment(horizontal="left", vertical="center")
            if fill:
                cell.fill = fill

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return _excel_response(output.read(), "tipos_solicitud.xlsx")


@login_required
def estado_solicitud_exportar_excel(request):
    """Exporta el listado de estados de solicitud a Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Estados de Solicitud"

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center")

    headers = ["Código", "Nombre", "Color", "Es Inicial", "Es Final", "Requiere Acción", "Activo"]
    widths  = [15,        35,       12,       12,           10,         18,                10]

    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"

    odd_fill = PatternFill("solid", fgColor="EBF0F8")
    for row_idx, obj in enumerate(
        EstadoSolicitud.objects.filter(eliminado=False).order_by('codigo'), start=2
    ):
        fill = odd_fill if row_idx % 2 == 0 else None
        color_hex = obj.color.lstrip('#') if hasattr(obj, 'color') and obj.color else "CCCCCC"
        fila = [
            obj.codigo,
            obj.nombre,
            obj.color if hasattr(obj, 'color') else "",
            "SI" if obj.es_inicial else "NO",
            "SI" if obj.es_final else "NO",
            "SI" if obj.requiere_accion else "NO",
            "SI" if obj.activo else "NO",
        ]
        for col, valor in enumerate(fila, start=1):
            cell = ws.cell(row=row_idx, column=col, value=valor)
            cell.alignment = center if col in (3, 4, 5, 6, 7) else Alignment(horizontal="left", vertical="center")
            if fill:
                cell.fill = fill
            # Colorear la celda de color con el color del estado
            if col == 3 and color_hex:
                try:
                    cell.fill = PatternFill("solid", fgColor=color_hex)
                except Exception:
                    pass

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return _excel_response(output.read(), "estados_solicitud.xlsx")


# ==================== EXPORTAR / IMPORTAR SOLICITUDES ====================

from .excel_services import ExportacionSolicitudesService, ImportacionSolicitudesService


@login_required
def solicitudes_exportar_excel(request):
    """Exporta la lista de solicitudes (con filtros) a Excel."""
    from .forms import FiltroSolicitudesForm

    queryset = Solicitud.objects.filter(eliminado=False)

    form = FiltroSolicitudesForm(request.GET)
    if form.is_valid():
        data = form.cleaned_data
        if data.get('estado'):
            queryset = queryset.filter(estado=data['estado'])
        if data.get('tipo'):
            queryset = queryset.filter(tipo_solicitud=data['tipo'])
        if data.get('fecha_desde'):
            queryset = queryset.filter(fecha_solicitud__gte=data['fecha_desde'])
        if data.get('fecha_hasta'):
            queryset = queryset.filter(fecha_solicitud__lte=data['fecha_hasta'])
        if data.get('buscar'):
            q = data['buscar']
            queryset = queryset.filter(
                Q(numero__icontains=q) |
                Q(solicitante__username__icontains=q) |
                Q(area__nombre__icontains=q) |
                Q(departamento__nombre__icontains=q)
            )

    contenido = ExportacionSolicitudesService.exportar(queryset, titulo="Reporte de Solicitudes")
    response = HttpResponse(
        contenido,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="solicitudes.xlsx"'
    return response


@login_required
def solicitudes_descargar_plantilla(request):
    """Descarga la plantilla Excel para importar solicitudes."""
    contenido = ImportacionSolicitudesService.generar_plantilla()
    response = HttpResponse(
        contenido,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_solicitudes.xlsx"'
    return response


@login_required
def solicitudes_importar_excel(request):
    """Importa solicitudes desde un archivo Excel."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Metodo no permitido'}, status=405)
    if 'archivo' not in request.FILES:
        return JsonResponse({'error': 'No se proporciono archivo'}, status=400)

    archivo = request.FILES['archivo']
    try:
        creadas, omitidas, errores = ImportacionSolicitudesService.importar(archivo, request.user)
        mensaje = f"Importacion completada: {creadas} solicitudes creadas"
        if omitidas:
            mensaje += f", {omitidas} omitidas"
        return JsonResponse({
            'success': True,
            'mensaje': mensaje,
            'creadas': creadas,
            'omitidas': omitidas,
            'errores': errores[:10],
        })
    except Exception as e:
        return JsonResponse({'error': f'Error al importar: {str(e)}'}, status=500)
