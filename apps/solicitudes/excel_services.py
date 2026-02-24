"""
Servicios de exportación e importación Excel para el módulo de solicitudes.

Provee:
- ExportacionSolicitudesService: exporta un queryset de Solicitud a Excel.
- ImportacionSolicitudesService: genera plantilla y permite importar solicitudes.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ── helpers de estilo ─────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_ODD_FILL = PatternFill("solid", fgColor="EBF0F8")
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


def _build_wb(titulo: str, headers_widths: list[tuple[str, int]], rows: list[list]) -> bytes:
    """Construye un workbook con encabezado estilizado y filas alternadas."""
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    for col, (h, w) in enumerate(headers_widths, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"

    for ri, fila in enumerate(rows, 2):
        fill = _ODD_FILL if ri % 2 == 0 else None
        for ci, v in enumerate(fila, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.alignment = _LEFT
            if fill:
                c.fill = fill

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


# ── Exportación ───────────────────────────────────────────────────────────────

class ExportacionSolicitudesService:
    """Exporta un queryset de Solicitud a un archivo Excel."""

    HEADERS_WIDTHS: list[tuple[str, int]] = [
        ("Número", 18),
        ("Fecha Solicitud", 20),
        ("Tipo", 12),
        ("Tipo Solicitud", 25),
        ("Estado", 22),
        ("Título Actividad", 35),
        ("Solicitante", 25),
        ("Departamento", 25),
        ("Área", 25),
        ("Motivo", 45),
        ("Observaciones", 45),
    ]

    @classmethod
    def exportar(cls, queryset, titulo: str = "Solicitudes") -> bytes:
        """Genera el Excel a partir del queryset de solicitudes."""
        from .models import Solicitud  # import local para evitar circulares

        qs = queryset.select_related(
            "tipo_solicitud", "estado", "solicitante", "departamento", "area"
        )

        rows = []
        for s in qs:
            rows.append([
                s.numero,
                s.fecha_solicitud.strftime("%d/%m/%Y %H:%M") if s.fecha_solicitud else "",
                s.get_tipo_display() if hasattr(s, "get_tipo_display") else s.tipo,
                s.tipo_solicitud.nombre if s.tipo_solicitud else "",
                s.estado.nombre if s.estado else "",
                s.titulo_actividad or "",
                s.solicitante.get_full_name() or s.solicitante.username if s.solicitante else "",
                s.departamento.nombre if s.departamento else "",
                s.area.nombre if s.area else "",
                s.motivo or "",
                s.observaciones or "",
            ])

        return _build_wb(titulo, cls.HEADERS_WIDTHS, rows)


# ── Importación ───────────────────────────────────────────────────────────────

class ImportacionSolicitudesService:
    """Genera plantilla e importa solicitudes desde Excel."""

    COLUMNAS = [
        "Numero", "TipoSolicitud", "Estado", "TituloActividad",
        "Departamento", "Area", "FechaRequerida", "Motivo", "Observaciones",
    ]

    @classmethod
    def generar_plantilla(cls) -> bytes:
        """Genera plantilla Excel con encabezados y hasta 5 solicitudes de ejemplo."""
        from .models import Solicitud

        wb = Workbook()
        ws = wb.active
        ws.title = "Solicitudes"

        for col_idx, enc in enumerate(cls.COLUMNAS, start=1):
            c = ws.cell(row=1, column=col_idx, value=enc)
            c.font = Font(bold=True)
            c.alignment = _CENTER

        for row_idx, obj in enumerate(
            Solicitud.objects.filter(eliminado=False)
            .select_related("tipo_solicitud", "estado", "departamento", "area")
            .order_by("-fecha_solicitud")[:5],
            start=2,
        ):
            ws.cell(row=row_idx, column=1, value=obj.numero)
            ws.cell(row=row_idx, column=2, value=obj.tipo_solicitud.codigo if obj.tipo_solicitud else "")
            ws.cell(row=row_idx, column=3, value=obj.estado.codigo if obj.estado else "")
            ws.cell(row=row_idx, column=4, value=obj.titulo_actividad or "")
            ws.cell(row=row_idx, column=5, value=obj.departamento.codigo if obj.departamento else "")
            ws.cell(row=row_idx, column=6, value=obj.area.codigo if obj.area else "")
            ws.cell(row=row_idx, column=7, value=obj.fecha_requerida.strftime("%d/%m/%Y") if obj.fecha_requerida else "")
            ws.cell(row=row_idx, column=8, value=obj.motivo or "")
            ws.cell(row=row_idx, column=9, value=obj.observaciones or "")

        col_widths = [18, 20, 20, 35, 22, 22, 16, 45, 45]
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        return contenido

    @classmethod
    def importar(cls, archivo, usuario) -> tuple[int, int, list[str]]:
        """
        Importa solicitudes desde Excel.

        Actualiza solicitudes existentes por número; omite filas con errores.
        Retorna (creadas, omitidas, errores).
        """
        from django.db import transaction
        from openpyxl import load_workbook
        from .models import Solicitud, TipoSolicitud, EstadoSolicitud, Departamento, Area
        import datetime

        wb = load_workbook(archivo, read_only=True)
        ws = wb.active

        encabezados = [str(c.value).strip() if c.value else "" for c in ws[1]]
        col_idx = {h: i for i, h in enumerate(encabezados)}

        def _get(fila, nombre: str) -> str:
            i = col_idx.get(nombre)
            if i is None:
                return ""
            v = fila[i].value
            return str(v).strip() if v is not None else ""

        creadas = 0
        omitidas = 0
        errores: list[str] = []

        with transaction.atomic():
            for row_num, fila in enumerate(ws.iter_rows(min_row=2), start=2):
                if all(c.value is None or str(c.value).strip() == "" for c in fila):
                    continue

                numero = _get(fila, "Numero")
                tipo_cod = _get(fila, "TipoSolicitud")
                estado_cod = _get(fila, "Estado")
                fecha_req_str = _get(fila, "FechaRequerida")
                motivo = _get(fila, "Motivo")

                if not numero or not tipo_cod or not estado_cod or not motivo:
                    errores.append(f"Fila {row_num}: Numero, TipoSolicitud, Estado y Motivo son obligatorios")
                    omitidas += 1
                    continue

                try:
                    tipo_solicitud = TipoSolicitud.objects.get(codigo=tipo_cod, eliminado=False)
                except TipoSolicitud.DoesNotExist:
                    errores.append(f"Fila {row_num}: TipoSolicitud '{tipo_cod}' no encontrado")
                    omitidas += 1
                    continue

                try:
                    estado = EstadoSolicitud.objects.get(codigo=estado_cod, eliminado=False)
                except EstadoSolicitud.DoesNotExist:
                    errores.append(f"Fila {row_num}: Estado '{estado_cod}' no encontrado")
                    omitidas += 1
                    continue

                dept_cod = _get(fila, "Departamento")
                area_cod = _get(fila, "Area")
                departamento = Departamento.objects.filter(codigo=dept_cod, eliminado=False).first() if dept_cod else None
                area = Area.objects.filter(codigo=area_cod, eliminado=False).first() if area_cod else None

                fecha_requerida = None
                if fecha_req_str:
                    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                        try:
                            fecha_requerida = datetime.datetime.strptime(fecha_req_str, fmt).date()
                            break
                        except ValueError:
                            continue

                try:
                    obj = Solicitud.objects.filter(numero=numero, eliminado=False).first()
                    if obj:
                        obj.tipo_solicitud = tipo_solicitud
                        obj.estado = estado
                        obj.titulo_actividad = _get(fila, "TituloActividad") or obj.titulo_actividad
                        obj.departamento = departamento
                        obj.area = area
                        if fecha_requerida:
                            obj.fecha_requerida = fecha_requerida
                        obj.motivo = motivo
                        obj.observaciones = _get(fila, "Observaciones") or obj.observaciones
                        obj.save()
                        creadas += 1
                    else:
                        omitidas += 1
                        errores.append(f"Fila {row_num}: Solicitud '{numero}' no existe (solo se permiten actualizaciones)")
                except Exception as e:
                    errores.append(f"Fila {row_num}: {e}")
                    omitidas += 1

        wb.close()
        return creadas, omitidas, errores
