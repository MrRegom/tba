"""
Vistas del módulo de bajas de inventario.

Este módulo implementa las vistas para la gestión de bajas de activos,
utilizando Class-Based Views y siguiendo las mejores prácticas de Django.

Todas las vistas usan CBVs y siguen type hints completos.
"""
from __future__ import annotations

from typing import Any

from django.db.models import QuerySet, Q
from django.urls import reverse_lazy
from django.views.generic import (
    TemplateView, ListView, DetailView, CreateView, UpdateView, DeleteView
)
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse

from core.mixins import (
    BaseAuditedViewMixin, AtomicTransactionMixin, SoftDeleteMixin,
    PaginatedListMixin
)
from .models import BajaInventario, MotivoBaja
from .forms import BajaInventarioForm, MotivoBajaForm, FiltroBajasForm


# ==================== VISTA MENÚ PRINCIPAL ====================

class MenuBajasView(BaseAuditedViewMixin, TemplateView):
    """
    Vista del menú principal del módulo de bajas de inventario.

    Muestra estadísticas y accesos rápidos basados en permisos del usuario.
    Permisos: bajas_inventario.view_bajainventario
    """
    template_name = 'bajas_inventario/menu_bajas.html'
    permission_required = 'bajas_inventario.view_bajainventario'

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        """Agrega estadísticas al contexto."""
        context = super().get_context_data(**kwargs)
        user = self.request.user

        # Estadísticas del módulo
        context['stats'] = {
            'total_bajas': BajaInventario.objects.filter(eliminado=False).count(),
            'mis_bajas': BajaInventario.objects.filter(
                solicitante=user, eliminado=False
            ).count(),
            'total_motivos': MotivoBaja.objects.filter(eliminado=False).count(),
        }

        # Permisos del usuario
        context['permisos'] = {
            'puede_crear': user.has_perm('bajas_inventario.add_bajainventario'),
            'puede_gestionar': user.has_perm('bajas_inventario.change_bajainventario'),
            'puede_motivos': user.has_perm('bajas_inventario.add_motivobaja'),
        }

        context['titulo'] = 'Módulo de Bajas de Inventario'
        return context


# ==================== VISTAS DE BAJAS DE INVENTARIO ====================

class BajaInventarioListView(BaseAuditedViewMixin, PaginatedListMixin, ListView):
    """
    Vista para listar bajas de inventario con filtros y paginación.

    Permisos: bajas_inventario.view_bajainventario
    Filtros: Motivo, fechas, búsqueda por texto
    """
    model = BajaInventario
    template_name = 'bajas_inventario/lista_bajas.html'
    context_object_name = 'bajas'
    permission_required = 'bajas_inventario.view_bajainventario'
    paginate_by = 25

    def get_queryset(self) -> QuerySet[BajaInventario]:
        """Retorna bajas con optimización N+1."""
        queryset = BajaInventario.objects.filter(eliminado=False).select_related(
            'motivo', 'ubicacion', 'solicitante', 'activo'
        )

        # Aplicar filtros del formulario
        form = FiltroBajasForm(self.request.GET)
        if form.is_valid():
            data = form.cleaned_data

            # Filtro por motivo
            if data.get('motivo'):
                queryset = queryset.filter(motivo=data['motivo'])

            # Filtro por fechas
            if data.get('fecha_desde'):
                queryset = queryset.filter(fecha_baja__gte=data['fecha_desde'])

            if data.get('fecha_hasta'):
                queryset = queryset.filter(fecha_baja__lte=data['fecha_hasta'])

            # Filtro de búsqueda por texto
            if data.get('buscar'):
                q = data['buscar']
                queryset = queryset.filter(
                    Q(numero__icontains=q) |
                    Q(activo__codigo__icontains=q) |
                    Q(activo__nombre__icontains=q) |
                    Q(observaciones__icontains=q)
                )

        return queryset.order_by('-fecha_baja', '-numero')

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        """Agrega datos adicionales al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Bajas de Inventario'
        context['form'] = FiltroBajasForm(self.request.GET)
        user = self.request.user
        context['permisos'] = {
            'puede_crear': user.has_perm('bajas_inventario.add_bajainventario'),
            'puede_editar': user.has_perm('bajas_inventario.change_bajainventario'),
            'puede_eliminar': user.has_perm('bajas_inventario.delete_bajainventario'),
        }
        return context


class BajaInventarioDetailView(BaseAuditedViewMixin, DetailView):
    """
    Vista para ver el detalle de una baja de inventario.

    Permisos: bajas_inventario.view_bajainventario
    """
    model = BajaInventario
    template_name = 'bajas_inventario/detalle_baja.html'
    context_object_name = 'baja'
    permission_required = 'bajas_inventario.view_bajainventario'

    def get_queryset(self) -> QuerySet[BajaInventario]:
        """Optimiza consultas con select_related."""
        return BajaInventario.objects.filter(eliminado=False).select_related(
            'motivo', 'ubicacion', 'solicitante', 'activo',
            'activo__categoria', 'activo__estado'
        )

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        """Agrega datos al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Baja {self.object.numero}'
        return context


class MisBajasListView(BaseAuditedViewMixin, PaginatedListMixin, ListView):
    """
    Vista para ver las bajas solicitadas por el usuario actual.

    Permisos: bajas_inventario.view_bajainventario
    """
    model = BajaInventario
    template_name = 'bajas_inventario/mis_bajas.html'
    context_object_name = 'bajas'
    permission_required = 'bajas_inventario.view_bajainventario'
    paginate_by = 25

    def get_queryset(self) -> QuerySet[BajaInventario]:
        """Retorna solo las bajas del usuario actual no eliminadas."""
        return BajaInventario.objects.filter(
            solicitante=self.request.user,
            eliminado=False
        ).select_related(
            'motivo', 'ubicacion', 'activo'
        ).order_by('-fecha_baja')

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        """Agrega datos adicionales al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Mis Bajas Solicitadas'
        return context


class BajaInventarioCreateView(BaseAuditedViewMixin, AtomicTransactionMixin, CreateView):
    """
    Vista para crear una nueva baja de inventario.

    Permisos: bajas_inventario.add_bajainventario
    Auditoría: Registra acción CREAR automáticamente
    Transacción atómica: Garantiza integridad de datos
    """
    model = BajaInventario
    form_class = BajaInventarioForm
    template_name = 'bajas_inventario/form_baja.html'
    permission_required = 'bajas_inventario.add_bajainventario'

    # Configuración de auditoría
    audit_action = 'CREAR'
    audit_description_template = 'Creó baja de inventario {obj.numero}'

    # Mensaje de éxito
    success_message = 'Baja {obj.numero} creada exitosamente.'

    def get_success_url(self) -> str:
        """Redirige al detalle de la baja creada."""
        return reverse_lazy('bajas_inventario:detalle_baja', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        """Agrega datos al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Nueva Baja de Inventario'
        context['action'] = 'Crear'
        return context

    def form_valid(self, form: BajaInventarioForm) -> HttpResponse:
        """Procesa el formulario válido asignando el solicitante."""
        baja = form.save(commit=False)
        baja.solicitante = self.request.user
        baja.save()
        self.object = baja

        response = super().form_valid(form)
        self.log_action(self.object, self.request)
        return response


class BajaInventarioUpdateView(BaseAuditedViewMixin, AtomicTransactionMixin, UpdateView):
    """
    Vista para editar una baja de inventario existente.

    Permisos: bajas_inventario.change_bajainventario
    Auditoría: Registra acción EDITAR automáticamente
    Transacción atómica: Garantiza integridad de datos
    """
    model = BajaInventario
    form_class = BajaInventarioForm
    template_name = 'bajas_inventario/form_baja.html'
    permission_required = 'bajas_inventario.change_bajainventario'

    # Configuración de auditoría
    audit_action = 'EDITAR'
    audit_description_template = 'Editó baja de inventario {obj.numero}'

    # Mensaje de éxito
    success_message = 'Baja {obj.numero} actualizada exitosamente.'

    def get_queryset(self) -> QuerySet[BajaInventario]:
        """Solo permite editar bajas no eliminadas."""
        return BajaInventario.objects.filter(eliminado=False)

    def get_success_url(self) -> str:
        """Redirige al detalle de la baja editada."""
        return reverse_lazy('bajas_inventario:detalle_baja', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        """Agrega datos al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Editar Baja {self.object.numero}'
        context['action'] = 'Editar'
        context['baja'] = self.object
        return context

    def form_valid(self, form: BajaInventarioForm) -> HttpResponse:
        """Procesa el formulario válido con log de auditoría."""
        response = super().form_valid(form)
        self.log_action(self.object, self.request)
        return response


class BajaInventarioDeleteView(BaseAuditedViewMixin, SoftDeleteMixin, DeleteView):
    """
    Vista para eliminar (soft delete) una baja de inventario.

    Permisos: bajas_inventario.delete_bajainventario
    Auditoría: Registra acción ELIMINAR automáticamente
    Implementa soft delete
    """
    model = BajaInventario
    template_name = 'bajas_inventario/eliminar_baja.html'
    permission_required = 'bajas_inventario.delete_bajainventario'
    success_url = reverse_lazy('bajas_inventario:lista_bajas')

    # Configuración de auditoría
    audit_action = 'ELIMINAR'
    audit_description_template = 'Eliminó baja de inventario {obj.numero}'

    # Mensaje de éxito
    success_message = 'Baja {obj.numero} eliminada exitosamente.'

    def get_queryset(self) -> QuerySet[BajaInventario]:
        """Solo permite eliminar bajas no eliminadas."""
        return BajaInventario.objects.filter(eliminado=False)

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        """Agrega datos al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Eliminar Baja {self.object.numero}'
        context['baja'] = self.object
        return context


# ==================== VISTAS DE MOTIVOS DE BAJA ====================

class MotivoBajaListView(BaseAuditedViewMixin, PaginatedListMixin, ListView):
    """Vista para listar motivos de baja."""
    model = MotivoBaja
    template_name = 'bajas_inventario/lista_motivos.html'
    context_object_name = 'motivos'
    permission_required = 'bajas_inventario.view_motivobaja'
    paginate_by = 25

    def get_queryset(self) -> QuerySet[MotivoBaja]:
        """Retorna solo motivos no eliminados."""
        return MotivoBaja.objects.filter(eliminado=False).order_by('codigo')

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        """Agrega datos adicionales al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Motivos de Baja'
        return context


class MotivoBajaCreateView(BaseAuditedViewMixin, CreateView):
    """Vista para crear un nuevo motivo de baja."""
    model = MotivoBaja
    form_class = MotivoBajaForm
    template_name = 'bajas_inventario/form_motivo.html'
    permission_required = 'bajas_inventario.add_motivobaja'
    success_url = reverse_lazy('bajas_inventario:lista_motivos')

    audit_action = 'CREAR'
    audit_description_template = 'Creó motivo de baja {obj.codigo} - {obj.nombre}'
    success_message = 'Motivo {obj.nombre} creado exitosamente.'

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        """Agrega datos al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Crear Motivo de Baja'
        context['action'] = 'Crear'
        return context

    def form_valid(self, form: MotivoBajaForm) -> HttpResponse:
        """Procesa el formulario válido con log de auditoría."""
        response = super().form_valid(form)
        self.log_action(self.object, self.request)
        return response


class MotivoBajaUpdateView(BaseAuditedViewMixin, UpdateView):
    """Vista para editar un motivo de baja existente."""
    model = MotivoBaja
    form_class = MotivoBajaForm
    template_name = 'bajas_inventario/form_motivo.html'
    permission_required = 'bajas_inventario.change_motivobaja'
    success_url = reverse_lazy('bajas_inventario:lista_motivos')

    audit_action = 'EDITAR'
    audit_description_template = 'Editó motivo de baja {obj.codigo} - {obj.nombre}'
    success_message = 'Motivo {obj.nombre} actualizado exitosamente.'

    def get_queryset(self) -> QuerySet[MotivoBaja]:
        """Solo permite editar motivos no eliminados."""
        return MotivoBaja.objects.filter(eliminado=False)

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        """Agrega datos al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Editar Motivo {self.object.codigo}'
        context['action'] = 'Editar'
        context['motivo'] = self.object
        return context

    def form_valid(self, form: MotivoBajaForm) -> HttpResponse:
        """Procesa el formulario válido con log de auditoría."""
        response = super().form_valid(form)
        self.log_action(self.object, self.request)
        return response


class MotivoBajaDeleteView(BaseAuditedViewMixin, SoftDeleteMixin, DeleteView):
    """Vista para eliminar (soft delete) un motivo de baja."""
    model = MotivoBaja
    template_name = 'bajas_inventario/eliminar_motivo.html'
    permission_required = 'bajas_inventario.delete_motivobaja'
    success_url = reverse_lazy('bajas_inventario:lista_motivos')

    audit_action = 'ELIMINAR'
    audit_description_template = 'Eliminó motivo de baja {obj.codigo} - {obj.nombre}'
    success_message = 'Motivo {obj.codigo} eliminado exitosamente.'

    def get_queryset(self) -> QuerySet[MotivoBaja]:
        """Solo permite eliminar motivos no eliminados."""
        return MotivoBaja.objects.filter(eliminado=False)

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        """Agrega datos al contexto."""
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Eliminar Motivo {self.object.codigo}'
        context['motivo'] = self.object
        return context


# ==================== GESTORES DE BAJAS ====================

class GestoresBajasView(BaseAuditedViewMixin, TemplateView):
    """Vista de gestores de bajas de inventario con tabs para cada entidad."""
    template_name = 'bajas_inventario/gestores_bajas.html'
    permission_required = 'bajas_inventario.view_motivobaja'

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context['motivos'] = MotivoBaja.objects.filter(eliminado=False).order_by('codigo')
        return context


# ── Exportaciones Excel (Gestores Bajas) ─────────────────────────────────────
from openpyxl import Workbook as _BWWB
from openpyxl.styles import Font as _BFont, PatternFill as _BPF, Alignment as _BAlign
from openpyxl.utils import get_column_letter as _bgcl
from io import BytesIO as _BBytesIO


def _bajas_wb(titulo, headers_widths, rows_data):
    wb = _BWWB(); ws = wb.active; ws.title = titulo[:31]
    hf = _BPF("solid", fgColor="1F3864"); hfont = _BFont(bold=True, color="FFFFFF", size=10)
    center = _BAlign(horizontal="center", vertical="center"); left = _BAlign(horizontal="left", vertical="center")
    odd = _BPF("solid", fgColor="EBF0F8")
    for col, (h, w) in enumerate(headers_widths, 1):
        c = ws.cell(row=1, column=col, value=h); c.font = hfont; c.fill = hf; c.alignment = center
        ws.column_dimensions[_bgcl(col)].width = w
    ws.freeze_panes = "A2"
    for ri, fila in enumerate(rows_data, 2):
        f = odd if ri % 2 == 0 else None
        for ci, v in enumerate(fila, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            if f: c.fill = f
            c.alignment = center if ci >= len(fila) else left
    out = _BBytesIO(); wb.save(out); out.seek(0); return out.read()


def _bajas_excel_resp(data, filename):
    resp = HttpResponse(data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'; return resp


@login_required
def motivo_baja_exportar_excel(request):
    rows = [(m.codigo, m.nombre, m.descripcion or "")
            for m in MotivoBaja.objects.filter(eliminado=False).order_by("codigo")]
    hdrs = [("Código", 15), ("Nombre", 35), ("Descripción", 50)]
    return _bajas_excel_resp(_bajas_wb("Motivos de Baja", hdrs, rows), "motivos_baja.xlsx")


@login_required
def motivo_baja_descargar_plantilla(request):
    from apps.bodega.excel_services.importacion_excel import ImportacionExcelService
    contenido = ImportacionExcelService.generar_plantilla_motivos_baja()
    response = HttpResponse(contenido, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="plantilla_motivos_baja.xlsx"'
    return response


@login_required
def motivo_baja_importar_excel(request):
    from apps.bodega.excel_services.importacion_excel import ImportacionExcelService
    if request.method != "POST":
        return JsonResponse({"success": False, "mensaje": "Método no permitido"}, status=405)
    archivo = request.FILES.get("archivo")
    valido, error = ImportacionExcelService.validar_archivo_excel(archivo)
    if not valido:
        return JsonResponse({"success": False, "mensaje": error})
    try:
        creadas, actualizadas, errores = ImportacionExcelService.importar_motivos_baja(archivo, request.user)
        return JsonResponse({
            "success": True,
            "mensaje": f"Importación completada: {creadas} creados, {actualizadas} actualizados.",
            "creadas": creadas,
            "actualizadas": actualizadas,
            "errores": errores[:10],
        })
    except Exception as e:
        return JsonResponse({"success": False, "mensaje": str(e)})
